
import os
import ntpath
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from sklearn.base import clone
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import (
    average_precision_score,
    precision_recall_curve,
    roc_auc_score,
    roc_curve,
    accuracy_score,
    balanced_accuracy_score,
    recall_score,
    precision_score,
    f1_score,
    confusion_matrix,
    matthews_corrcoef,
)

from imblearn.pipeline import Pipeline as ImbPipeline
from imblearn.over_sampling import SMOTE, BorderlineSMOTE, SVMSMOTE
from catboost import CatBoostClassifier

warnings.filterwarnings("ignore")


# =============================================================================
# 0. Basic settings: mainly edit here
# =============================================================================
DATA_FILE = r"D:\萝卜投稿专用\Original Research - Spinal Infection Following Internal Fixation Surgery\插值选择\Xgboost\lr=0.1\训练集-z-scores.xlsx"

# 外部验证集：必须是使用训练集 scaler transform 后的 z-score 文件，且列名与训练集最终指标一致。
EXTERNAL_DATA_FILE = r"D:\萝卜投稿专用\Original Research - Spinal Infection Following Internal Fixation Surgery\插值选择\Xgboost\lr=0.1\外部验证集的处理\External validation set_filtered by training set items.xlsx"

TARGET_COL = "Infection"
POS_LABEL = 1

RANDOM_STATE = 42
N_SPLITS = 5
THRESHOLD = 0.30

MODEL_SHORT = "CatBoost"

# Root output folder. A timestamped subfolder will be automatically created inside it.
OUTPUT_ROOT = r"D:\萝卜投稿专用\Original Research - Spinal Infection Following Internal Fixation Surgery\插值选择\Xgboost\lr=0.1\外部验证集的处理"

# 当前输入已经是 z-score 标准化后的数据，所以这里保持 False。
USE_SCALER = False
SMOTE_K_NEIGHBORS = 5

# 已经确定的 SMOTE 比例；如仍想比较多个比例，可写成 [0.4, 0.65] 等。
SAMPLING_RATIOS = [0.41]

MODEL_PARAMS = {
    "iterations": 100,
    "learning_rate": 0.085,
    "depth": 1,
    "l2_leaf_reg": 10,
    "random_strength": 11,
    "bagging_temperature": 8,
    "loss_function": "Logloss",
    "eval_metric": "Logloss",
    "random_seed": RANDOM_STATE,
    "verbose": False,
    "allow_writing_files": False,
}

SAMPLER_CONFIGS = {
    
    "BorderlineSMOTE": BorderlineSMOTE
   
}

# 内部 5-fold fold-level bootstrap 数；只用于内部选择过程中的曲线标注辅助，不输出表。
BOOTSTRAP_N = 2000
BOOTSTRAP_CI = 95
BOOTSTRAP_RANDOM_STATE = 2026

# 选择最佳设置时仍然沿用原来的规则。
TARGET_VAL_AUPRC_LOW = 0.70
TARGET_VAL_AUPRC_HIGH = 0.80
MAX_GAP = 0.10


# =============================================================================
# 1. Plot style
# =============================================================================
plt.rcParams["font.family"] = ["Arial", "DejaVu Sans"]
plt.rcParams["font.size"] = 10
plt.rcParams["axes.linewidth"] = 0.8
plt.rcParams["pdf.fonttype"] = 42
plt.rcParams["ps.fonttype"] = 42

COLOR_TRAIN = "#4E79A7"
COLOR_VAL = "#E15759"
COLOR_EXTERNAL = "#59A14F"


# =============================================================================
# 2. Utilities
# =============================================================================
def ensure_output_dir(path):
    if path:
        os.makedirs(path, exist_ok=True)


def make_timestamped_output_dir():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = os.path.join(OUTPUT_ROOT, f"CatBoost_EXTERNAL_{timestamp}")
    ensure_output_dir(out_dir)
    return out_dir


def resolve_data_file(file_path):
    """Try original path first, then current folder and script folder by basename."""
    if os.path.exists(file_path):
        return file_path

    base_names = []
    base_names.append(os.path.basename(file_path))
    base_names.append(ntpath.basename(file_path))
    base_names = [b for b in dict.fromkeys(base_names) if b]

    search_dirs = [os.getcwd()]
    try:
        search_dirs.append(os.path.dirname(os.path.abspath(__file__)))
    except Exception:
        pass

    for d in search_dirs:
        for b in base_names:
            candidate = os.path.join(d, b)
            if os.path.exists(candidate):
                return candidate

    raise FileNotFoundError(
        f"找不到数据文件：{file_path}\n"
        f"请确认Excel文件与本py在同一文件夹，或把路径改成真实路径。"
    )


def load_data(file_path, target_col):
    file_path = resolve_data_file(file_path)
    df = pd.read_excel(file_path)

    if target_col not in df.columns:
        raise ValueError(f"找不到结局变量列：{target_col}；当前列名：{list(df.columns)}")

    df = df.dropna(axis=1, how="all")
    df = df.dropna(subset=[target_col]).copy()

    y = pd.to_numeric(df[target_col], errors="coerce")
    X = df.drop(columns=[target_col])

    non_numeric_cols = X.select_dtypes(exclude=[np.number]).columns.tolist()
    if non_numeric_cols:
        print("以下非数值列将被自动删除：")
        print(non_numeric_cols)
        X = X.drop(columns=non_numeric_cols)

    if X.isna().sum().sum() > 0:
        print("警告：训练集X中存在缺失值，将用训练集列中位数填补。")
        X = X.fillna(X.median(numeric_only=True))

    valid = y.notna()
    X = X.loc[valid].reset_index(drop=True)
    y = y.loc[valid].astype(int).reset_index(drop=True)
    y = (y == POS_LABEL).astype(int)

    if y.nunique() != 2:
        raise ValueError("结局变量必须包含0和1两类。请检查 Infection 列。")

    return X, y


def load_external_data(file_path, target_col, feature_columns, train_medians):
    file_path = resolve_data_file(file_path)
    df = pd.read_excel(file_path)

    if target_col not in df.columns:
        raise ValueError(f"外部验证集找不到结局变量列：{target_col}；当前列名：{list(df.columns)}")

    missing_cols = [c for c in feature_columns if c not in df.columns]
    if missing_cols:
        raise ValueError(
            "外部验证集缺少训练集最终模型所需指标：\n"
            + "\n".join(missing_cols)
            + "\n请先按训练集最终指标处理外部验证集。"
        )

    y_ext = pd.to_numeric(df[target_col], errors="coerce")
    X_ext = df.loc[:, feature_columns].copy()

    for col in X_ext.columns:
        X_ext[col] = pd.to_numeric(X_ext[col], errors="coerce")

    if X_ext.isna().sum().sum() > 0:
        print("警告：外部验证集X中存在缺失值，将用训练集同名列中位数填补。")
        X_ext = X_ext.fillna(train_medians)

    valid = y_ext.notna()
    X_ext = X_ext.loc[valid].reset_index(drop=True)
    y_ext = y_ext.loc[valid].astype(int).reset_index(drop=True)
    y_ext = (y_ext == POS_LABEL).astype(int)

    if y_ext.nunique() != 2:
        print("警告：外部验证集结局不是0/1两类均存在，AUC/PR曲线可能无法计算。")

    return X_ext, y_ext


def build_model():
    return CatBoostClassifier(**MODEL_PARAMS)


def safe_k_neighbors(y, requested_k=SMOTE_K_NEIGHBORS):
    minority_n = int(min(np.sum(y == 0), np.sum(y == 1)))
    return max(1, min(requested_k, minority_n - 1))


def make_sampler(sampler_name, sampler_class, ratio, y=None):
    k = SMOTE_K_NEIGHBORS if y is None else safe_k_neighbors(np.asarray(y), requested_k=SMOTE_K_NEIGHBORS)
    kwargs = {
        "sampling_strategy": ratio,
        "random_state": RANDOM_STATE,
        "k_neighbors": k,
    }
    if sampler_name == "BorderlineSMOTE":
        return sampler_class(**kwargs, kind="borderline-1")
    if sampler_name == "SVMSMOTE":
        return sampler_class(**kwargs, m_neighbors=min(10, max(1, k)))
    return sampler_class(**kwargs)


def make_pipeline(sampler, model):
    steps = []
    if USE_SCALER:
        from sklearn.preprocessing import StandardScaler
        steps.append(("scaler", StandardScaler()))
    steps.append(("smote", sampler))
    steps.append(("model", model))
    return ImbPipeline(steps=steps)


def get_positive_probability(model, X_data):
    if hasattr(model, "predict_proba"):
        return model.predict_proba(X_data)[:, 1]
    if hasattr(model, "decision_function"):
        score = model.decision_function(X_data)
        return 1 / (1 + np.exp(-score))
    return model.predict(X_data)


def calc_metrics(y_true, y_prob, threshold=THRESHOLD):
    y_true = np.asarray(y_true).astype(int)
    y_prob = np.asarray(y_prob).astype(float)
    y_pred = (y_prob >= threshold).astype(int)

    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    specificity = tn / (tn + fp) if (tn + fp) > 0 else np.nan

    if len(np.unique(y_true)) == 2:
        auc_value = roc_auc_score(y_true, y_prob)
        auprc_value = average_precision_score(y_true, y_prob)
    else:
        auc_value = np.nan
        auprc_value = np.nan

    return {
        "AUPRC": auprc_value,
        "AUC": auc_value,
        "Accuracy": accuracy_score(y_true, y_pred),
        "Balanced_Accuracy": balanced_accuracy_score(y_true, y_pred),
        "Recall_Sensitivity": recall_score(y_true, y_pred, zero_division=0),
        "Specificity": specificity,
        "Precision": precision_score(y_true, y_pred, zero_division=0),
        "F1": f1_score(y_true, y_pred, zero_division=0),
        "MCC": matthews_corrcoef(y_true, y_pred),
        "TP": tp,
        "FP": fp,
        "TN": tn,
        "FN": fn,
    }


def summarize_fold_metrics(fold_df):
    ignore_cols = ["Fold", "Dataset", "SMOTE_Method", "Sampling_Ratio"]
    metrics = [c for c in fold_df.columns if c not in ignore_cols]
    rows = []
    for dataset in ["Train", "Validation"]:
        sub = fold_df[fold_df["Dataset"] == dataset]
        row = {"Dataset": dataset}
        for m in metrics:
            row[m + "_Mean"] = float(sub[m].mean())
            row[m + "_SD"] = float(sub[m].std(ddof=1)) if len(sub) > 1 else 0.0
        rows.append(row)
    return pd.DataFrame(rows)


def get_mean_metric(summary_df, dataset, metric):
    return float(summary_df.loc[summary_df["Dataset"] == dataset, f"{metric}_Mean"].iloc[0])


def save_excel_formatted(df, path, sheet_name="Sheet1"):
    ensure_output_dir(os.path.dirname(path))
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        ws = writer.sheets[sheet_name[:31]]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="center")
                cell.border = border
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 70)


def sci_axes(ax):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(axis="both", direction="out", length=3, width=0.8)
    ax.grid(False)


def save_figure(fig, out_dir, name):
    ensure_output_dir(out_dir)
    fig.savefig(os.path.join(out_dir, f"{name}.png"), dpi=600, bbox_inches="tight")
    fig.savefig(os.path.join(out_dir, f"{name}.pdf"), bbox_inches="tight")
    plt.close(fig)


def percent_ci_format(mean, lower, upper):
    if pd.isna(mean):
        return "NA"
    return f"{100 * mean:.2f}% ({100 * lower:.2f}–{100 * upper:.2f}%)"


def metric_format(mean, lower, upper, metric):
    if metric == "MCC":
        if pd.isna(mean):
            return "NA"
        return f"{mean:.3f} ({lower:.3f}–{upper:.3f})"
    return percent_ci_format(mean, lower, upper)


# =============================================================================
# 3. CV and selection
# =============================================================================
def run_one_setting(X, y, sampler_name, sampler_class, ratio):
    cv = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)
    fold_rows = []
    val_y_all, val_prob_all = [], []
    train_y_all, train_prob_all = [], []

    for fold_id, (train_idx, val_idx) in enumerate(cv.split(X, y), start=1):
        X_train, X_val = X.iloc[train_idx], X.iloc[val_idx]
        y_train, y_val = y.iloc[train_idx], y.iloc[val_idx]

        # SMOTE 只发生在这一折的训练部分，不会作用到 validation fold。
        sampler = make_sampler(sampler_name, sampler_class, ratio, y_train)
        pipeline = make_pipeline(sampler, clone(build_model()))
        pipeline.fit(X_train, y_train)

        train_prob = get_positive_probability(pipeline, X_train)
        val_prob = get_positive_probability(pipeline, X_val)

        for dataset, yy, pp in [
            ("Train", y_train, train_prob),
            ("Validation", y_val, val_prob),
        ]:
            row = {
                "Fold": fold_id,
                "Dataset": dataset,
                "SMOTE_Method": sampler_name,
                "Sampling_Ratio": ratio,
            }
            row.update(calc_metrics(yy, pp, THRESHOLD))
            fold_rows.append(row)

        train_y_all.extend(y_train.tolist())
        train_prob_all.extend(train_prob.tolist())
        val_y_all.extend(y_val.tolist())
        val_prob_all.extend(val_prob.tolist())

    fold_df = pd.DataFrame(fold_rows)
    summary = summarize_fold_metrics(fold_df)
    val_auprc = get_mean_metric(summary, "Validation", "AUPRC")
    train_auprc = get_mean_metric(summary, "Train", "AUPRC")

    return {
        "Model": MODEL_SHORT,
        "SMOTE_Method": sampler_name,
        "Sampling_Ratio": ratio,
        "Train_AUPRC": train_auprc,
        "Val_AUPRC": val_auprc,
        "Gap": train_auprc - val_auprc,
        "Threshold": THRESHOLD,
        "Params": str(MODEL_PARAMS),
        "Fold_DF": fold_df,
        "Summary_DF": summary,
        "train_y": np.array(train_y_all),
        "train_prob": np.array(train_prob_all),
        "val_y": np.array(val_y_all),
        "val_prob": np.array(val_prob_all),
    }


def add_selection_flags(result_df):
    df = result_df.copy()
    df["In_Target_PR_Range"] = (df["Val_AUPRC"] >= TARGET_VAL_AUPRC_LOW) & (df["Val_AUPRC"] <= TARGET_VAL_AUPRC_HIGH)
    df["Gap_OK"] = df["Gap"].abs() <= MAX_GAP
    df["Eligible"] = df["In_Target_PR_Range"] & df["Gap_OK"]
    return df


def train_final_model_on_full_training(X, y, sampler_name, sampler_class, ratio):
    # 最终模型：完整训练集做 SMOTE，然后训练一个最终模型；外部验证集不参与训练。
    sampler = make_sampler(sampler_name, sampler_class, ratio, y)
    pipeline = make_pipeline(sampler, build_model())
    pipeline.fit(X, y)
    return pipeline


# =============================================================================
# 4. External bootstrap and plots
# =============================================================================
def bootstrap_external_metrics(y_true, y_prob, output_dir):
    rng = np.random.default_rng(BOOTSTRAP_RANDOM_STATE)
    y_true = np.asarray(y_true).astype(int)
    y_prob = np.asarray(y_prob).astype(float)
    n = len(y_true)

    metric_cols = [
        "AUPRC", "AUC", "Accuracy", "Balanced_Accuracy", "Recall_Sensitivity",
        "Specificity", "Precision", "F1", "MCC",
    ]

    base = calc_metrics(y_true, y_prob, THRESHOLD)
    boot_values = {m: [] for m in metric_cols}

    for _ in range(BOOTSTRAP_N):
        idx = rng.choice(np.arange(n), size=n, replace=True)
        yy = y_true[idx]
        pp = y_prob[idx]
        try:
            row = calc_metrics(yy, pp, THRESHOLD)
            for m in metric_cols:
                boot_values[m].append(row[m])
        except Exception:
            for m in metric_cols:
                boot_values[m].append(np.nan)

    alpha = (100 - BOOTSTRAP_CI) / 2
    records = []
    for m in metric_cols:
        arr = np.asarray(boot_values[m], dtype=float)
        arr = arr[~np.isnan(arr)]
        if len(arr) == 0:
            lower, upper = np.nan, np.nan
        else:
            lower, upper = np.nanpercentile(arr, [alpha, 100 - alpha])
        mean_value = float(base[m]) if m in base else np.nan
        records.append({
            "Dataset": "External_Test",
            "Metric": m,
            "Value": mean_value,
            "CI_Lower": float(lower) if not pd.isna(lower) else np.nan,
            "CI_Upper": float(upper) if not pd.isna(upper) else np.nan,
            "Formatted_95CI": metric_format(mean_value, lower, upper, m),
            "N_External": n,
            "Bootstrap_N": BOOTSTRAP_N,
            "Threshold": THRESHOLD,
        })

    # 混淆矩阵计数只报告原始外部验证集结果，不做百分比CI。
    count_records = []
    for m in ["TP", "FP", "TN", "FN"]:
        count_records.append({
            "Dataset": "External_Test",
            "Metric": m,
            "Value": int(base[m]),
            "CI_Lower": "",
            "CI_Upper": "",
            "Formatted_95CI": str(int(base[m])),
            "N_External": n,
            "Bootstrap_N": "",
            "Threshold": THRESHOLD,
        })

    out_df = pd.DataFrame(records + count_records)
    save_excel_formatted(out_df, os.path.join(output_dir, "01_external_test_metrics_bootstrap_95CI.xlsx"), "External_95CI")
    return out_df, base


def plot_train_val_external_curves(best_obj, best_row, best_summary, y_ext, prob_ext, external_metrics_df, output_dir):
    fig_dir = os.path.join(output_dir, "PR_ROC_curves_WITH_EXTERNAL")
    ensure_output_dir(fig_dir)

    best_method = str(best_row["SMOTE_Method"])
    best_ratio = float(best_row["Sampling_Ratio"])
    ratio_tag = str(best_ratio).replace(".", "p")

    train_auprc_mean = float(best_row["Train_AUPRC"])
    val_auprc_mean = float(best_row["Val_AUPRC"])
    train_auc_mean = get_mean_metric(best_summary, "Train", "AUC")
    val_auc_mean = get_mean_metric(best_summary, "Validation", "AUC")

    metrics_index = external_metrics_df.set_index("Metric")
    ext_auprc_label = metrics_index.loc["AUPRC", "Formatted_95CI"] if "AUPRC" in metrics_index.index else "NA"
    ext_auc_label = metrics_index.loc["AUC", "Formatted_95CI"] if "AUC" in metrics_index.index else "NA"

    # PR curve: Train / Validation 来自 pooled 5-fold CV，External 来自最终完整训练集模型。
    train_precision, train_recall, _ = precision_recall_curve(best_obj["train_y"], best_obj["train_prob"])
    val_precision, val_recall, _ = precision_recall_curve(best_obj["val_y"], best_obj["val_prob"])
    ext_precision, ext_recall, _ = precision_recall_curve(y_ext, prob_ext)

    fig, ax = plt.subplots(figsize=(6.4, 5.2), dpi=600)
    ax.plot(train_recall, train_precision, linewidth=1.8, color=COLOR_TRAIN,
            label=f"Train mean AUPRC = {train_auprc_mean:.3f}")
    ax.plot(val_recall, val_precision, linewidth=1.8, color=COLOR_VAL,
            label=f"Validation mean AUPRC = {val_auprc_mean:.3f}")
    ax.plot(ext_recall, ext_precision, linewidth=1.8, color=COLOR_EXTERNAL,
            label=f"External AUPRC = {ext_auprc_label}")
    ax.set_xlabel("Recall")
    ax.set_ylabel("Precision")
    ax.set_title(f"{MODEL_SHORT} Precision-Recall Curve")
    ax.legend(frameon=False, loc="best", fontsize=8)
    sci_axes(ax)
    fig.tight_layout()
    pr_name = f"PR_curve_Train_Validation_External_BEST_{best_method}_ratio_{ratio_tag}".replace(".", "p")
    save_figure(fig, fig_dir, pr_name)

    # ROC curve
    train_fpr, train_tpr, _ = roc_curve(best_obj["train_y"], best_obj["train_prob"])
    val_fpr, val_tpr, _ = roc_curve(best_obj["val_y"], best_obj["val_prob"])

    fig, ax = plt.subplots(figsize=(6.4, 5.2), dpi=600)
    ax.plot(train_fpr, train_tpr, linewidth=1.8, color=COLOR_TRAIN,
            label=f"Train mean AUC = {train_auc_mean:.3f}")
    ax.plot(val_fpr, val_tpr, linewidth=1.8, color=COLOR_VAL,
            label=f"Validation mean AUC = {val_auc_mean:.3f}")

    if len(np.unique(y_ext)) == 2:
        ext_fpr, ext_tpr, _ = roc_curve(y_ext, prob_ext)
        ax.plot(ext_fpr, ext_tpr, linewidth=1.8, color=COLOR_EXTERNAL,
                label=f"External AUC = {ext_auc_label}")
    else:
        ax.plot([], [], linewidth=1.8, color=COLOR_EXTERNAL,
                label="External AUC = NA")

    ax.plot([0, 1], [0, 1], linestyle="--", linewidth=1.0, color="gray")
    ax.set_xlabel("False Positive Rate")
    ax.set_ylabel("True Positive Rate")
    ax.set_title(f"{MODEL_SHORT} ROC Curve")
    ax.legend(frameon=False, loc="best", fontsize=8)
    sci_axes(ax)
    fig.tight_layout()
    roc_name = f"ROC_curve_Train_Validation_External_BEST_{best_method}_ratio_{ratio_tag}".replace(".", "p")
    save_figure(fig, fig_dir, roc_name)


# =============================================================================
# 5. Main
# =============================================================================
def main():
    OUTPUT_DIR = make_timestamped_output_dir()
    X, y = load_data(DATA_FILE, TARGET_COL)
    train_medians = X.median(numeric_only=True)
    X_ext, y_ext = load_external_data(EXTERNAL_DATA_FILE, TARGET_COL, list(X.columns), train_medians)

    print("=" * 80)
    print("RUNNING CATBOOST WITH EXTERNAL VALIDATION")
    print(f"Output folder: {OUTPUT_DIR}")
    print(f"Training N={len(y)}, Positive={int(y.sum())}, Negative={int(len(y) - y.sum())}")
    print(f"External N={len(y_ext)}, Positive={int(y_ext.sum())}, Negative={int(len(y_ext) - y_ext.sum())}")
    print(f"Model params={MODEL_PARAMS}")
    print(f"SMOTE k_neighbors={SMOTE_K_NEIGHBORS}")
    print(f"Threshold={THRESHOLD}")
    print("=" * 80)

    all_results = []
    summaries = {}
    full_objs = {}

    current_ratio = y.sum() / (len(y) - y.sum())

    for sampler_name, sampler_class in SAMPLER_CONFIGS.items():
        for ratio in SAMPLING_RATIOS:
            if ratio <= current_ratio:
                print(f"Skip {sampler_name} ratio={ratio}: ratio <= current minority/majority ratio")
                continue

            print(f"Running internal CV: {sampler_name}, ratio={ratio}")
            obj = run_one_setting(X, y, sampler_name, sampler_class, ratio)

            all_results.append({
                "Model": obj["Model"],
                "SMOTE_Method": obj["SMOTE_Method"],
                "Sampling_Ratio": obj["Sampling_Ratio"],
                "Train_AUPRC": obj["Train_AUPRC"],
                "Val_AUPRC": obj["Val_AUPRC"],
                "Gap": obj["Gap"],
                "Threshold": obj["Threshold"],
                "Params": obj["Params"],
            })
            key = f"{sampler_name}_{ratio}"
            summaries[key] = obj["Summary_DF"]
            full_objs[key] = obj

    if len(all_results) == 0:
        raise RuntimeError("没有任何SMOTE设置被运行。请检查 SAMPLING_RATIOS 是否大于当前少数类/多数类比例。")

    result_df = pd.DataFrame(all_results)
    result_df = add_selection_flags(result_df)
    result_df = result_df.sort_values(["Val_AUPRC", "Gap"], ascending=[False, True]).reset_index(drop=True)

    best = result_df.iloc[0]
    best_sampler_name = best["SMOTE_Method"]
    best_ratio = float(best["Sampling_Ratio"])
    best_key = f"{best_sampler_name}_{best_ratio}"
    best_sampler_class = SAMPLER_CONFIGS[best_sampler_name]
    best_obj = full_objs[best_key]
    best_summary = summaries[best_key]

    # 最终模型：完整训练集 + 最佳SMOTE设置。
    final_model = train_final_model_on_full_training(
        X, y, best_sampler_name, best_sampler_class, best_ratio
    )
    ext_prob = get_positive_probability(final_model, X_ext)

    # 只输出外部验证集指标及 bootstrap 95%CI。
    external_metrics_df, external_base = bootstrap_external_metrics(y_ext, ext_prob, OUTPUT_DIR)

    # 保存外部预测概率。
    pred_df = pd.DataFrame({
        "y_true": y_ext,
        "y_prob": ext_prob,
        "y_pred": (ext_prob >= THRESHOLD).astype(int),
    })
    save_excel_formatted(pred_df, os.path.join(OUTPUT_DIR, "02_external_test_predictions.xlsx"), "External_Pred")

    # 保存最终模型设置信息，不输出train/validation各项指标表。
    setting_df = pd.DataFrame([{
        "Model": MODEL_SHORT,
        "Best_SMOTE_Method": best_sampler_name,
        "Best_Sampling_Ratio": best_ratio,
        "Threshold": THRESHOLD,
        "N_Train": len(y),
        "Positive_Train": int(y.sum()),
        "N_External": len(y_ext),
        "Positive_External": int(y_ext.sum()),
        "Params": str(MODEL_PARAMS),
        "Note": "Best setting selected by internal 5-fold CV. Final model trained on full training set with selected SMOTE. External test was not used for selection.",
    }])
    save_excel_formatted(setting_df, os.path.join(OUTPUT_DIR, "00_final_model_setting.xlsx"), "Setting")

    # 三条线同图：Train / Validation / External。
    plot_train_val_external_curves(best_obj, best, best_summary, y_ext, ext_prob, external_metrics_df, OUTPUT_DIR)

    print("=" * 80)
    print("FINISHED.")
    print(f"IMPORTANT: Open this NEW output folder only:\n{OUTPUT_DIR}")
    print("External metrics:")
    print(external_metrics_df.to_string(index=False))
    print("=" * 80)


if __name__ == "__main__":
    main()
