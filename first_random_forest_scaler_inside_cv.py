
import os
import ntpath
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from sklearn.base import clone
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import (
    average_precision_score,
    precision_recall_curve,
    roc_auc_score,
    roc_curve,
    accuracy_score,
    balanced_accuracy_score,
    recall_score,
    precision_score,
    f1_score,
    confusion_matrix,
    matthews_corrcoef,
)
from sklearn.feature_selection import RFECV
from sklearn.ensemble import RandomForestClassifier

from imblearn.pipeline import Pipeline as ImbPipeline
from imblearn.over_sampling import SMOTE, BorderlineSMOTE, SVMSMOTE

warnings.filterwarnings("ignore")


# 基本设置：主要改这里
DATA_FILE = r"D:\萝卜投稿专用\Original Research - Spinal Infection Following Internal Fixation Surgery\插值选择\Xgboost\lr=0.1\P小于0.05_按临床动态标准共线性删除后.xlsx"
# NOTE: For leakage-free standardization, DATA_FILE should be the imputed/raw training data, not a pre-z-scored file.
TARGET_COL = "Infection"
POS_LABEL = 1

RANDOM_STATE = 42
N_SPLITS = 5
THRESHOLD = 0.3
TOP_N_FEATURES = 10

OUTPUT_DIR = r"D:\萝卜投稿专用\Original Research - Spinal Infection Following Internal Fixation Surgery\插值选择\Xgboost\lr=0.1\建模-1_RF_k15_BOOTSTRAP"
MODEL_SHORT = "Random Forest"

USE_SCALER = True   # FIXED: scaler is fitted within each CV training fold; use raw/unstandardized input data.
SMOTE_K_NEIGHBORS = 17

# Bootstrap设置：用于最终表格的95%CI，不参与模型训练/筛选
BOOTSTRAP_N = 1000
BOOTSTRAP_RANDOM_STATE = 2026
BOOTSTRAP_CI = 0.95

SAMPLING_RATIOS = [0.46]

MODEL_PARAMS = {
    "n_estimators": 67,
    "max_depth": 1,
    "min_samples_leaf": 24,
    "min_samples_split": 24,
    "max_features": "log2",
    "class_weight": None,
    "random_state": RANDOM_STATE,
    "n_jobs": -1,
}

SAMPLER_CONFIGS = {
    "SMOTE": SMOTE,
    "BorderlineSMOTE": BorderlineSMOTE, 
    "SVMSMOTE": SVMSMOTE

}


# 作图风格
plt.rcParams["font.family"] = ["Arial", "DejaVu Sans"]
plt.rcParams["font.size"] = 10
plt.rcParams["axes.linewidth"] = 0.8
plt.rcParams["pdf.fonttype"] = 42
plt.rcParams["ps.fonttype"] = 42

COLOR_TRAIN = "#4E79A7"
COLOR_VAL = "#E15759"
COLOR_BAR = "#4E79A7"


def ensure_output_dir(path):
    if path:
        os.makedirs(path, exist_ok=True)


def resolve_data_file(file_path):
    if os.path.exists(file_path):
        return file_path

    base_names = []
    base_names.append(os.path.basename(file_path))
    base_names.append(ntpath.basename(file_path))  # 处理Windows反斜杠路径
    base_names = [b for b in dict.fromkeys(base_names) if b]

    search_dirs = [os.getcwd()]
    try:
        search_dirs.append(os.path.dirname(os.path.abspath(__file__)))
    except Exception:
        pass

    for d in search_dirs:
        for b in base_names:
            candidate = os.path.join(d, b)
            if os.path.exists(candidate):
                return candidate

    raise FileNotFoundError(
        f"找不到数据文件：{file_path}\n"
        f"请确认Excel文件与本py在同一文件夹，或把 DATA_FILE 改成真实路径。"
    )


def load_data(file_path, target_col):
    file_path = resolve_data_file(file_path)
    df = pd.read_excel(file_path)

    if target_col not in df.columns:
        raise ValueError(f"找不到结局变量列：{target_col}；当前列名：{list(df.columns)}")

    df = df.dropna(axis=1, how="all")
    df = df.dropna(subset=[target_col]).copy()

    y = pd.to_numeric(df[target_col], errors="coerce")
    X = df.drop(columns=[target_col])

    non_numeric_cols = X.select_dtypes(exclude=[np.number]).columns.tolist()
    if non_numeric_cols:
        print("以下非数值列将被自动删除：")
        print(non_numeric_cols)
        X = X.drop(columns=non_numeric_cols)

    if X.isna().sum().sum() > 0:
        print("警告：X中存在缺失值，将用列中位数填补。")
        X = X.fillna(X.median(numeric_only=True))

    valid = y.notna()
    X = X.loc[valid].reset_index(drop=True)
    y = y.loc[valid].astype(int).reset_index(drop=True)
    y = (y == POS_LABEL).astype(int)

    if y.nunique() != 2:
        raise ValueError("结局变量必须包含0和1两类。请检查 Infection 列。")

    return X, y


def build_model():
    return RandomForestClassifier(**MODEL_PARAMS)


def safe_k_neighbors(y, requested_k=SMOTE_K_NEIGHBORS):
    minority_n = int(min(np.sum(y == 0), np.sum(y == 1)))
    return max(1, min(requested_k, minority_n - 1))


def make_sampler(sampler_name, sampler_class, ratio, y=None):
    k = SMOTE_K_NEIGHBORS if y is None else safe_k_neighbors(np.asarray(y), requested_k=SMOTE_K_NEIGHBORS)
    kwargs = {
        "sampling_strategy": ratio,
        "random_state": RANDOM_STATE,
        "k_neighbors": k,
    }
    if sampler_name == "BorderlineSMOTE":
        return sampler_class(**kwargs, kind="borderline-1")
    if sampler_name == "SVMSMOTE":
        return sampler_class(**kwargs, m_neighbors=min(10, max(1, k)))
    return sampler_class(**kwargs)


def make_pipeline(sampler, model):
    steps = []
    if USE_SCALER:
        from sklearn.preprocessing import StandardScaler
        steps.append(("scaler", StandardScaler()))
    steps.append(("smote", sampler))
    steps.append(("model", model))
    return ImbPipeline(steps=steps)


def get_positive_probability(model, X_data):
    if hasattr(model, "predict_proba"):
        return model.predict_proba(X_data)[:, 1]
    if hasattr(model, "decision_function"):
        score = model.decision_function(X_data)
        return 1 / (1 + np.exp(-score))
    return model.predict(X_data)


def calc_metrics(y_true, y_prob, threshold=0.5):
    y_pred = (y_prob >= threshold).astype(int)
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    specificity = tn / (tn + fp) if (tn + fp) > 0 else np.nan

    return {
        "AUPRC": average_precision_score(y_true, y_prob),
        "AUC": roc_auc_score(y_true, y_prob),
        "Accuracy": accuracy_score(y_true, y_pred),
        "Balanced_Accuracy": balanced_accuracy_score(y_true, y_pred),
        "Recall_Sensitivity": recall_score(y_true, y_pred, zero_division=0),
        "Specificity": specificity,
        "Precision": precision_score(y_true, y_pred, zero_division=0),
        "F1": f1_score(y_true, y_pred, zero_division=0),
        "MCC": matthews_corrcoef(y_true, y_pred),
        "TP": tp,
        "FP": fp,
        "TN": tn,
        "FN": fn,
    }


def summarize_fold_metrics(fold_df):
    ignore_cols = ["Fold", "Dataset", "SMOTE_Method", "Sampling_Ratio"]
    metrics = [c for c in fold_df.columns if c not in ignore_cols]
    rows = []
    for dataset in ["Train", "Validation"]:
        sub = fold_df[fold_df["Dataset"] == dataset]
        row = {"Dataset": dataset}
        for m in metrics:
            row[m + "_Mean"] = float(sub[m].mean())
            row[m + "_SD"] = float(sub[m].std(ddof=1)) if len(sub) > 1 else 0.0
        rows.append(row)
    return pd.DataFrame(rows)


def save_excel_formatted(df, path, sheet_name="Sheet1"):
    ensure_output_dir(os.path.dirname(path))
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        ws = writer.sheets[sheet_name[:31]]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="center")
                cell.border = border
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)


def sci_axes(ax):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(axis="both", direction="out", length=3, width=0.8)
    ax.grid(False)


def save_figure(fig, out_dir, name):
    ensure_output_dir(out_dir)
    fig.savefig(os.path.join(out_dir, f"{name}.png"), dpi=600, bbox_inches="tight")
    fig.savefig(os.path.join(out_dir, f"{name}.pdf"), bbox_inches="tight")
    plt.close(fig)


def get_mean_metric(summary_df, dataset, metric):
    return float(summary_df.loc[summary_df["Dataset"] == dataset, f"{metric}_Mean"].iloc[0])


def plot_combined_pr_roc(train_y, train_prob, val_y, val_prob, summary_df, output_dir):

    fig_dir = os.path.join(output_dir, "PR_ROC_curves")
    ensure_output_dir(fig_dir)

    train_auprc_mean = get_mean_metric(summary_df, "Train", "AUPRC")
    val_auprc_mean = get_mean_metric(summary_df, "Validation", "AUPRC")
    train_auc_mean = get_mean_metric(summary_df, "Train", "AUC")
    val_auc_mean = get_mean_metric(summary_df, "Validation", "AUC")

    # PR曲线
    train_precision, train_recall, _ = precision_recall_curve(train_y, train_prob)
    val_precision, val_recall, _ = precision_recall_curve(val_y, val_prob)

    fig, ax = plt.subplots(figsize=(6, 5), dpi=600)
    ax.plot(train_recall, train_precision, linewidth=1.8, color=COLOR_TRAIN,
            label=f"Train mean AUPRC = {train_auprc_mean:.3f}")
    ax.plot(val_recall, val_precision, linewidth=1.8, color=COLOR_VAL,
            label=f"Validation mean AUPRC = {val_auprc_mean:.3f}")
    ax.set_xlabel("Recall")
    ax.set_ylabel("Precision")
    ax.set_title(f"{MODEL_SHORT} Precision-Recall Curve")
    ax.legend(frameon=False, loc="best")
    sci_axes(ax)
    fig.tight_layout()
    save_figure(fig, fig_dir, "Train_Validation_PR_curve")

    # ROC曲线
    train_fpr, train_tpr, _ = roc_curve(train_y, train_prob)
    val_fpr, val_tpr, _ = roc_curve(val_y, val_prob)

    fig, ax = plt.subplots(figsize=(6, 5), dpi=600)
    ax.plot(train_fpr, train_tpr, linewidth=1.8, color=COLOR_TRAIN,
            label=f"Train mean AUC = {train_auc_mean:.3f}")
    ax.plot(val_fpr, val_tpr, linewidth=1.8, color=COLOR_VAL,
            label=f"Validation mean AUC = {val_auc_mean:.3f}")
    ax.plot([0, 1], [0, 1], linestyle="--", linewidth=1.0, color="gray")
    ax.set_xlabel("False Positive Rate")
    ax.set_ylabel("True Positive Rate")
    ax.set_title(f"{MODEL_SHORT} ROC Curve")
    ax.legend(frameon=False, loc="best")
    sci_axes(ax)
    fig.tight_layout()
    save_figure(fig, fig_dir, "Train_Validation_ROC_curve")


def run_one_setting(X, y, sampler_name, sampler_class, ratio):
    cv = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)
    fold_rows = []
    val_y_all, val_prob_all = [], []
    train_y_all, train_prob_all = [], []

    for fold_id, (train_idx, val_idx) in enumerate(cv.split(X, y), start=1):
        X_train, X_val = X.iloc[train_idx], X.iloc[val_idx]
        y_train, y_val = y.iloc[train_idx], y.iloc[val_idx]

        sampler = make_sampler(sampler_name, sampler_class, ratio, y_train)
        pipeline = make_pipeline(sampler, clone(build_model()))
        pipeline.fit(X_train, y_train)

        train_prob = get_positive_probability(pipeline, X_train)
        val_prob = get_positive_probability(pipeline, X_val)

        for dataset, yy, pp in [
            ("Train", y_train, train_prob),
            ("Validation", y_val, val_prob),
        ]:
            row = {
                "Fold": fold_id,
                "Dataset": dataset,
                "SMOTE_Method": sampler_name,
                "Sampling_Ratio": ratio,
            }
            row.update(calc_metrics(yy, pp, THRESHOLD))
            fold_rows.append(row)

        train_y_all.extend(y_train.tolist())
        train_prob_all.extend(train_prob.tolist())
        val_y_all.extend(y_val.tolist())
        val_prob_all.extend(val_prob.tolist())

    fold_df = pd.DataFrame(fold_rows)
    summary = summarize_fold_metrics(fold_df)
    val_auprc = get_mean_metric(summary, "Validation", "AUPRC")
    train_auprc = get_mean_metric(summary, "Train", "AUPRC")

    return {
        "Model": "RandomForest",
        "SMOTE_Method": sampler_name,
        "Sampling_Ratio": ratio,
        "Train_AUPRC": train_auprc,
        "Val_AUPRC": val_auprc,
        "Gap": train_auprc - val_auprc,
        "Params": str(MODEL_PARAMS),
        "Fold_DF": fold_df,
        "Summary_DF": summary,
        "train_y": np.array(train_y_all),
        "train_prob": np.array(train_prob_all),
        "val_y": np.array(val_y_all),
        "val_prob": np.array(val_prob_all),
    }


def run_rfecv_and_plot_top10(X, y, sampler_name, sampler_class, ratio):
    rfecv_dir = os.path.join(OUTPUT_DIR, "RFECV")
    ensure_output_dir(rfecv_dir)

    sampler = make_sampler(sampler_name, sampler_class, ratio, y)
    pipeline = make_pipeline(sampler, build_model())
    cv = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)

    rfecv = RFECV(
        estimator=pipeline,
        step=1,
        cv=cv,
        scoring="average_precision",
        n_jobs=-1,
        importance_getter="named_steps.model.feature_importances_",
        min_features_to_select=1,
    )
    rfecv.fit(X, y)

    ranking = pd.DataFrame({
        "Feature": X.columns,
        "Selected_by_RFECV": rfecv.support_,
        "RFECV_Ranking": rfecv.ranking_,
    }).sort_values(["RFECV_Ranking", "Feature"], ascending=[True, True]).reset_index(drop=True)

    selected_features = X.columns[rfecv.support_].tolist()
    selected_df = pd.DataFrame({"Selected_Feature": selected_features})

    save_excel_formatted(selected_df, os.path.join(rfecv_dir, "RFECV_selected_features.xlsx"), "Selected")
    save_excel_formatted(ranking, os.path.join(rfecv_dir, "RFECV_feature_ranking.xlsx"), "RFECV_Ranking")

    cv_scores = pd.DataFrame({
        "Number_of_Selected_Features": np.arange(1, len(rfecv.cv_results_["mean_test_score"]) + 1),
        "Mean_CV_AUPRC": rfecv.cv_results_["mean_test_score"],
        "SD_CV_AUPRC": rfecv.cv_results_.get("std_test_score", np.nan),
    })
    save_excel_formatted(cv_scores, os.path.join(rfecv_dir, "RFECV_cv_scores.xlsx"), "CV_Scores")

    # RFECV筛选后，只在筛选出的特征上重新训练模型并画Top10重要性
    X_selected = X[selected_features]
    X_selected_model = fit_transform_full_training_if_needed(X_selected, rfecv_dir, prefix="rfecv_selected_full_training")
    final_sampler = make_sampler(sampler_name, sampler_class, ratio, y)
    X_res, y_res = final_sampler.fit_resample(X_selected_model, y)

    final_model = build_model()
    final_model.fit(X_res, y_res)

    rfecv_imp = pd.DataFrame({
        "Feature": selected_features,
        "Importance_after_RFECV": final_model.feature_importances_,
    }).sort_values("Importance_after_RFECV", ascending=False).reset_index(drop=True)

    save_excel_formatted(
        rfecv_imp,
        os.path.join(rfecv_dir, "RFECV_selected_feature_importance.xlsx"),
        "RFECV_Importance",
    )

    plot_df = rfecv_imp.head(TOP_N_FEATURES).sort_values("Importance_after_RFECV", ascending=True)
    height = max(4.5, 0.38 * len(plot_df))
    fig, ax = plt.subplots(figsize=(8, height), dpi=600)
    ax.barh(
        plot_df["Feature"],
        plot_df["Importance_after_RFECV"],
        color=COLOR_BAR,
        edgecolor="black",
        linewidth=0.4,
    )
    ax.set_xlabel("Feature importance after RFECV")
    ax.set_ylabel("")
    ax.set_title(f"Top {min(TOP_N_FEATURES, len(plot_df))} RFECV-Selected Features")
    sci_axes(ax)
    fig.tight_layout()
    save_figure(fig, rfecv_dir, "RFECV_top10_selected_features")

    return ranking, rfecv_imp


def bootstrap_ci_from_fold_metrics(
    fold_df,
    n_bootstrap=BOOTSTRAP_N,
    random_state=BOOTSTRAP_RANDOM_STATE,
    ci=BOOTSTRAP_CI,
):
    """
    Bootstrap一般放在模型交叉验证完成之后：
    1）先用5折CV得到每一折的Train/Validation指标；
    2）再对“折”进行有放回抽样；
    3）得到每个指标的mean和95%CI。

    这个函数不重新训练模型，不影响模型选择，只用于最终表格的置信区间。
    """
    metric_cols = [
        "AUPRC",
        "AUC",
        "Accuracy",
        "Balanced_Accuracy",
        "Recall_Sensitivity",
        "Specificity",
        "Precision",
        "F1",
        "MCC",
    ]
    metric_cols = [c for c in metric_cols if c in fold_df.columns]

    rng = np.random.default_rng(random_state)
    alpha = 1.0 - ci
    lower_q = 100 * alpha / 2
    upper_q = 100 * (1 - alpha / 2)

    rows = []
    for dataset in ["Train", "Validation"]:
        sub = fold_df[fold_df["Dataset"] == dataset].copy()
        n = len(sub)
        if n == 0:
            continue

        for metric in metric_cols:
            values = sub[metric].astype(float).to_numpy()
            point_mean = float(np.mean(values))
            point_sd = float(np.std(values, ddof=1)) if len(values) > 1 else 0.0

            boot_means = []
            for _ in range(n_bootstrap):
                sample = rng.choice(values, size=n, replace=True)
                boot_means.append(float(np.mean(sample)))

            ci_lower = float(np.percentile(boot_means, lower_q))
            ci_upper = float(np.percentile(boot_means, upper_q))

            rows.append({
                "Dataset": dataset,
                "Metric": metric,
                "Mean": point_mean,
                "SD": point_sd,
                "CI_Lower": ci_lower,
                "CI_Upper": ci_upper,
                "Bootstrap_N": n_bootstrap,
                "CI_Level": ci,
                "Bootstrap_Unit": "5-fold metrics",
            })

    return pd.DataFrame(rows)


def percent_ci_text(mean, low, high):
    return f"{mean * 100:.2f}% ({low * 100:.2f}–{high * 100:.2f}%)"


def decimal_ci_text(mean, low, high):
    return f"{mean:.3f} ({low:.3f}–{high:.3f})"


def make_paper_table_row_from_bootstrap(bootstrap_df, model_name, sampler_name, ratio):
    """生成可以直接放进论文Table的单行：默认用Validation结果。"""
    sub = bootstrap_df[bootstrap_df["Dataset"] == "Validation"].copy()
    metric_map = {row["Metric"]: row for _, row in sub.iterrows()}

    def get_percent(metric):
        if metric not in metric_map:
            return ""
        r = metric_map[metric]
        return percent_ci_text(r["Mean"], r["CI_Lower"], r["CI_Upper"])

    return pd.DataFrame([{
        "Model": model_name,
        "SMOTE_Method": sampler_name,
        "Sampling_Ratio": ratio,
        "Balanced accuracy": get_percent("Balanced_Accuracy"),
        "Recall": get_percent("Recall_Sensitivity"),
        "Precision": get_percent("Precision"),
        "F1 score": get_percent("F1"),
        "AUPRC": get_percent("AUPRC"),
        "AUROC": get_percent("AUC"),
        "Specificity": get_percent("Specificity"),
        "Accuracy": get_percent("Accuracy"),
    }])


def warn_if_prezscore_input(file_path):
    """Warn if the input file name suggests that features were standardized before CV."""
    name = os.path.basename(str(file_path)).lower()
    if any(k in name for k in ["z-score", "z_scores", "z-scores", "zscore", "standard"]):
        print("警告：当前 DATA_FILE 文件名看起来像已经整体标准化过的数据。")
        print("为了完全避免标准化泄漏，建议改用未标准化的插补后原始训练集。")
        print("本脚本已把 StandardScaler 放进CV流程：每一折只在训练折fit，再transform验证折。")


def fit_transform_full_training_if_needed(X, output_dir=None, prefix="full_training"):
    """
    For post-CV final refitting / interpretation only.
    The scaler is fitted on the entire development cohort after model selection.
    Do NOT use external validation data here.
    """
    if not USE_SCALER:
        return X.copy()
    from sklearn.preprocessing import StandardScaler
    scaler = StandardScaler()
    X_scaled = pd.DataFrame(scaler.fit_transform(X), columns=X.columns, index=X.index)
    if output_dir is not None:
        try:
            scaler_df = pd.DataFrame({
                "Feature": X.columns,
                "Training_Mean": scaler.mean_,
                "Training_SD": scaler.scale_,
            })
            save_excel_formatted(
                scaler_df,
                os.path.join(output_dir, f"{prefix}_scaler_mean_sd.xlsx"),
                "Scaler_Mean_SD",
            )
        except Exception as e:
            print(f"Scaler mean/SD保存失败: {e}")
    return X_scaled



def main():
    ensure_output_dir(OUTPUT_DIR)
    X, y = load_data(DATA_FILE, TARGET_COL)
    warn_if_prezscore_input(DATA_FILE)

    print("=" * 70)
    print(f"N={len(y)}, Positive={int(y.sum())}, Negative={int(len(y) - y.sum())}")
    print(f"Positive ratio={y.mean():.4f}")
    print(f"Model params={MODEL_PARAMS}")
    print(f"SMOTE k_neighbors={SMOTE_K_NEIGHBORS}")
    print("=" * 70)

    all_results = []
    all_folds = []
    summaries = {}
    full_objs = {}

    current_ratio = y.sum() / (len(y) - y.sum())

    for sampler_name, sampler_class in SAMPLER_CONFIGS.items():
        for ratio in SAMPLING_RATIOS:
            if ratio <= current_ratio:
                print(f"Skip {sampler_name} ratio={ratio}: ratio <= current minority/majority ratio")
                continue

            print(f"Running: {sampler_name}, ratio={ratio}")
            obj = run_one_setting(X, y, sampler_name, sampler_class, ratio)

            all_results.append({
                "Model": obj["Model"],
                "SMOTE_Method": obj["SMOTE_Method"],
                "Sampling_Ratio": obj["Sampling_Ratio"],
                "Train_AUPRC": obj["Train_AUPRC"],
                "Val_AUPRC": obj["Val_AUPRC"],
                "Gap": obj["Gap"],
                "Params": obj["Params"],
            })
            all_folds.append(obj["Fold_DF"])
            key = f"{sampler_name}_{ratio}"
            summaries[key] = obj["Summary_DF"]
            full_objs[key] = obj

    if len(all_results) == 0:
        raise RuntimeError("没有任何SMOTE设置被运行。请检查 SAMPLING_RATIOS 是否大于当前少数类/多数类比例。")

    result_df = pd.DataFrame(all_results).sort_values(
        ["Val_AUPRC", "Gap"], ascending=[False, True]
    ).reset_index(drop=True)

    best = result_df.iloc[0]
    best_sampler_name = best["SMOTE_Method"]
    best_ratio = float(best["Sampling_Ratio"])
    best_key = f"{best_sampler_name}_{best_ratio}"
    best_sampler_class = SAMPLER_CONFIGS[best_sampler_name]
    best_obj = full_objs[best_key]
    best_summary = summaries[best_key]

    # Bootstrap放在这里：模型和最佳SMOTE已确定之后，只给最终指标补95%CI
    bootstrap_ci_df = bootstrap_ci_from_fold_metrics(best_obj["Fold_DF"])
    paper_table_df = make_paper_table_row_from_bootstrap(
        bootstrap_ci_df,
        model_name="Random Forest",
        sampler_name=best_sampler_name,
        ratio=best_ratio,
    )

    # 只输出必要文件
    save_excel_formatted(result_df, os.path.join(OUTPUT_DIR, "01_all_results.xlsx"), "All_Results")
    save_excel_formatted(pd.concat(all_folds, ignore_index=True), os.path.join(OUTPUT_DIR, "02_fold_metrics.xlsx"), "Fold_Metrics")
    save_excel_formatted(best_summary, os.path.join(OUTPUT_DIR, "03_best_summary_metrics_mean_sd.xlsx"), "Summary")
    save_excel_formatted(result_df.head(1), os.path.join(OUTPUT_DIR, "04_best_setting.xlsx"), "Best_Setting")
    save_excel_formatted(bootstrap_ci_df, os.path.join(OUTPUT_DIR, "05_best_fold_bootstrap_95CI.xlsx"), "Bootstrap_95CI")
    save_excel_formatted(paper_table_df, os.path.join(OUTPUT_DIR, "06_paper_table_row_validation_95CI.xlsx"), "Paper_Table")

    # 合并PR/ROC图：图例使用5折均值，和Excel一致
    plot_combined_pr_roc(
        best_obj["train_y"],
        best_obj["train_prob"],
        best_obj["val_y"],
        best_obj["val_prob"],
        best_summary,
        OUTPUT_DIR,
    )

    # 只生成RFECV筛选后的Top10特征重要性图
    run_rfecv_and_plot_top10(X, y, best_sampler_name, best_sampler_class, best_ratio)

    print("=" * 70)
    print("Finished. Results saved to:", OUTPUT_DIR)
    print("Best setting:")
    print(result_df.head(1).to_string(index=False))
    print("Best summary:")
    print(best_summary.to_string(index=False))
    print("Paper table row with bootstrap 95% CI:")
    print(paper_table_df.to_string(index=False))
    print("=" * 70)


if __name__ == "__main__":
    main()
