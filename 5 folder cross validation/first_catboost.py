
import os
import ntpath
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from sklearn.base import clone
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import (
    average_precision_score,
    precision_recall_curve,
    roc_auc_score,
    roc_curve,
    accuracy_score,
    balanced_accuracy_score,
    recall_score,
    precision_score,
    f1_score,
    confusion_matrix,
    matthews_corrcoef,
)

from imblearn.pipeline import Pipeline as ImbPipeline
from imblearn.over_sampling import SMOTE, BorderlineSMOTE, SVMSMOTE
from catboost import CatBoostClassifier

warnings.filterwarnings("ignore")


# =============================================================================
# 0. Basic settings: mainly edit here
# =============================================================================
DATA_FILE = r"D:\萝卜投稿专用\Original Research - Spinal Infection Following Internal Fixation Surgery\插值选择\Xgboost\lr=0.1\训练集-z-scores.xlsx"
TARGET_COL = "Infection"
POS_LABEL = 1

RANDOM_STATE = 42
N_SPLITS = 5
THRESHOLD = 0.30
TOP_N_FEATURES = 10

MODEL_SHORT = "CatBoost"

# Root output folder. A timestamped subfolder will be automatically created inside it.
OUTPUT_ROOT = r"D:\萝卜投稿专用\Original Research - Spinal Infection Following Internal Fixation Surgery\插值选择\Xgboost\lr=0.1\建模-1_catboost_FINAL_CLEAN"

USE_SCALER = False
SMOTE_K_NEIGHBORS = 5

SAMPLING_RATIOS = [0.41]

MODEL_PARAMS = {
    "iterations": 100,
    "learning_rate": 0.085,
    "depth": 1,
    "l2_leaf_reg": 10,
    "random_strength": 11,
    "bagging_temperature": 8,
    "loss_function": "Logloss",
    "eval_metric": "Logloss",
    "random_seed": RANDOM_STATE,
    "verbose": False,
    "allow_writing_files": False,
}

SAMPLER_CONFIGS = {
    "SMOTE": SMOTE,
    "BorderlineSMOTE": BorderlineSMOTE,
    "SVMSMOTE": SVMSMOTE,
}

# Bootstrap over 5 fold-level metrics, after the best setting has been selected.
BOOTSTRAP_N = 2000
BOOTSTRAP_CI = 95
BOOTSTRAP_RANDOM_STATE = 2026

TARGET_VAL_AUPRC_LOW = 0.70
TARGET_VAL_AUPRC_HIGH = 0.80
MAX_GAP = 0.10


# =============================================================================
# 1. Plot style
# =============================================================================
plt.rcParams["font.family"] = ["Arial", "DejaVu Sans"]
plt.rcParams["font.size"] = 10
plt.rcParams["axes.linewidth"] = 0.8
plt.rcParams["pdf.fonttype"] = 42
plt.rcParams["ps.fonttype"] = 42

COLOR_TRAIN = "#4E79A7"
COLOR_VAL = "#E15759"
COLOR_BAR = "#4E79A7"


# =============================================================================
# 2. Utilities
# =============================================================================
def ensure_output_dir(path):
    if path:
        os.makedirs(path, exist_ok=True)


def make_timestamped_output_dir():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = os.path.join(OUTPUT_ROOT, f"CatBoost_FINAL_CLEAN_{timestamp}")
    ensure_output_dir(out_dir)
    return out_dir


def resolve_data_file(file_path):
    """Try original path first, then current folder and script folder by basename."""
    if os.path.exists(file_path):
        return file_path

    base_names = []
    base_names.append(os.path.basename(file_path))
    base_names.append(ntpath.basename(file_path))
    base_names = [b for b in dict.fromkeys(base_names) if b]

    search_dirs = [os.getcwd()]
    try:
        search_dirs.append(os.path.dirname(os.path.abspath(__file__)))
    except Exception:
        pass

    for d in search_dirs:
        for b in base_names:
            candidate = os.path.join(d, b)
            if os.path.exists(candidate):
                return candidate

    raise FileNotFoundError(
        f"找不到数据文件：{file_path}\n"
        f"请确认Excel文件与本py在同一文件夹，或把 DATA_FILE 改成真实路径。"
    )


def load_data(file_path, target_col):
    file_path = resolve_data_file(file_path)
    df = pd.read_excel(file_path)

    if target_col not in df.columns:
        raise ValueError(f"找不到结局变量列：{target_col}；当前列名：{list(df.columns)}")

    df = df.dropna(axis=1, how="all")
    df = df.dropna(subset=[target_col]).copy()

    y = pd.to_numeric(df[target_col], errors="coerce")
    X = df.drop(columns=[target_col])

    non_numeric_cols = X.select_dtypes(exclude=[np.number]).columns.tolist()
    if non_numeric_cols:
        print("以下非数值列将被自动删除：")
        print(non_numeric_cols)
        X = X.drop(columns=non_numeric_cols)

    if X.isna().sum().sum() > 0:
        print("警告：X中存在缺失值，将用列中位数填补。")
        X = X.fillna(X.median(numeric_only=True))

    valid = y.notna()
    X = X.loc[valid].reset_index(drop=True)
    y = y.loc[valid].astype(int).reset_index(drop=True)
    y = (y == POS_LABEL).astype(int)

    if y.nunique() != 2:
        raise ValueError("结局变量必须包含0和1两类。请检查 Infection 列。")

    return X, y


def build_model():
    return CatBoostClassifier(**MODEL_PARAMS)


def safe_k_neighbors(y, requested_k=SMOTE_K_NEIGHBORS):
    minority_n = int(min(np.sum(y == 0), np.sum(y == 1)))
    return max(1, min(requested_k, minority_n - 1))


def make_sampler(sampler_name, sampler_class, ratio, y=None):
    k = SMOTE_K_NEIGHBORS if y is None else safe_k_neighbors(np.asarray(y), requested_k=SMOTE_K_NEIGHBORS)
    kwargs = {
        "sampling_strategy": ratio,
        "random_state": RANDOM_STATE,
        "k_neighbors": k,
    }
    if sampler_name == "BorderlineSMOTE":
        return sampler_class(**kwargs, kind="borderline-1")
    if sampler_name == "SVMSMOTE":
        return sampler_class(**kwargs, m_neighbors=min(10, max(1, k)))
    return sampler_class(**kwargs)


def make_pipeline(sampler, model):
    steps = []
    if USE_SCALER:
        from sklearn.preprocessing import StandardScaler
        steps.append(("scaler", StandardScaler()))
    steps.append(("smote", sampler))
    steps.append(("model", model))
    return ImbPipeline(steps=steps)


def get_positive_probability(model, X_data):
    if hasattr(model, "predict_proba"):
        return model.predict_proba(X_data)[:, 1]
    if hasattr(model, "decision_function"):
        score = model.decision_function(X_data)
        return 1 / (1 + np.exp(-score))
    return model.predict(X_data)


def calc_metrics(y_true, y_prob, threshold=THRESHOLD):
    y_pred = (y_prob >= threshold).astype(int)
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    specificity = tn / (tn + fp) if (tn + fp) > 0 else np.nan

    return {
        "AUPRC": average_precision_score(y_true, y_prob),
        "AUC": roc_auc_score(y_true, y_prob),
        "Accuracy": accuracy_score(y_true, y_pred),
        "Balanced_Accuracy": balanced_accuracy_score(y_true, y_pred),
        "Recall_Sensitivity": recall_score(y_true, y_pred, zero_division=0),
        "Specificity": specificity,
        "Precision": precision_score(y_true, y_pred, zero_division=0),
        "F1": f1_score(y_true, y_pred, zero_division=0),
        "MCC": matthews_corrcoef(y_true, y_pred),
        "TP": tp,
        "FP": fp,
        "TN": tn,
        "FN": fn,
    }


def summarize_fold_metrics(fold_df):
    ignore_cols = ["Fold", "Dataset", "SMOTE_Method", "Sampling_Ratio"]
    metrics = [c for c in fold_df.columns if c not in ignore_cols]
    rows = []
    for dataset in ["Train", "Validation"]:
        sub = fold_df[fold_df["Dataset"] == dataset]
        row = {"Dataset": dataset}
        for m in metrics:
            row[m + "_Mean"] = float(sub[m].mean())
            row[m + "_SD"] = float(sub[m].std(ddof=1)) if len(sub) > 1 else 0.0
        rows.append(row)
    return pd.DataFrame(rows)


def get_mean_metric(summary_df, dataset, metric):
    return float(summary_df.loc[summary_df["Dataset"] == dataset, f"{metric}_Mean"].iloc[0])


def save_excel_formatted(df, path, sheet_name="Sheet1"):
    ensure_output_dir(os.path.dirname(path))
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        ws = writer.sheets[sheet_name[:31]]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="center")
                cell.border = border
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def sci_axes(ax):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(axis="both", direction="out", length=3, width=0.8)
    ax.grid(False)


def save_figure(fig, out_dir, name):
    ensure_output_dir(out_dir)
    fig.savefig(os.path.join(out_dir, f"{name}.png"), dpi=600, bbox_inches="tight")
    fig.savefig(os.path.join(out_dir, f"{name}.pdf"), bbox_inches="tight")
    plt.close(fig)


def percent_ci_format(mean, lower, upper):
    return f"{100 * mean:.2f}% ({100 * lower:.2f}–{100 * upper:.2f}%)"


# =============================================================================
# 3. CV, selection, plotting, bootstrap
# =============================================================================
def run_one_setting(X, y, sampler_name, sampler_class, ratio):
    cv = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)
    fold_rows = []
    val_y_all, val_prob_all = [], []
    train_y_all, train_prob_all = [], []

    for fold_id, (train_idx, val_idx) in enumerate(cv.split(X, y), start=1):
        X_train, X_val = X.iloc[train_idx], X.iloc[val_idx]
        y_train, y_val = y.iloc[train_idx], y.iloc[val_idx]

        sampler = make_sampler(sampler_name, sampler_class, ratio, y_train)
        pipeline = make_pipeline(sampler, clone(build_model()))
        pipeline.fit(X_train, y_train)

        train_prob = get_positive_probability(pipeline, X_train)
        val_prob = get_positive_probability(pipeline, X_val)

        for dataset, yy, pp in [
            ("Train", y_train, train_prob),
            ("Validation", y_val, val_prob),
        ]:
            row = {
                "Fold": fold_id,
                "Dataset": dataset,
                "SMOTE_Method": sampler_name,
                "Sampling_Ratio": ratio,
            }
            row.update(calc_metrics(yy, pp, THRESHOLD))
            fold_rows.append(row)

        train_y_all.extend(y_train.tolist())
        train_prob_all.extend(train_prob.tolist())
        val_y_all.extend(y_val.tolist())
        val_prob_all.extend(val_prob.tolist())

    fold_df = pd.DataFrame(fold_rows)
    summary = summarize_fold_metrics(fold_df)
    val_auprc = get_mean_metric(summary, "Validation", "AUPRC")
    train_auprc = get_mean_metric(summary, "Train", "AUPRC")

    return {
        "Model": MODEL_SHORT,
        "SMOTE_Method": sampler_name,
        "Sampling_Ratio": ratio,
        "Train_AUPRC": train_auprc,
        "Val_AUPRC": val_auprc,
        "Gap": train_auprc - val_auprc,
        "Threshold": THRESHOLD,
        "Params": str(MODEL_PARAMS),
        "Fold_DF": fold_df,
        "Summary_DF": summary,
        "train_y": np.array(train_y_all),
        "train_prob": np.array(train_prob_all),
        "val_y": np.array(val_y_all),
        "val_prob": np.array(val_prob_all),
    }


def add_selection_flags(result_df):
    df = result_df.copy()
    df["In_Target_PR_Range"] = (df["Val_AUPRC"] >= TARGET_VAL_AUPRC_LOW) & (df["Val_AUPRC"] <= TARGET_VAL_AUPRC_HIGH)
    df["Gap_OK"] = df["Gap"].abs() <= MAX_GAP
    df["Eligible"] = df["In_Target_PR_Range"] & df["Gap_OK"]
    return df


def plot_combined_pr_roc(best_obj, best_row, best_summary, output_dir):
    """
    Curve shape uses pooled CV predictions from the selected best setting.
    Legend values are forcibly taken from best_row / best_summary of the selected best setting.
    Therefore, the numbers must match 01_all_results.xlsx exactly.
    """
    fig_dir = os.path.join(output_dir, "PR_ROC_curves_FINAL")
    ensure_output_dir(fig_dir)

    best_method = str(best_row["SMOTE_Method"])
    best_ratio = float(best_row["Sampling_Ratio"])
    ratio_tag = str(best_ratio).replace(".", "p")

    train_auprc_mean = float(best_row["Train_AUPRC"])
    val_auprc_mean = float(best_row["Val_AUPRC"])
    train_auc_mean = get_mean_metric(best_summary, "Train", "AUC")
    val_auc_mean = get_mean_metric(best_summary, "Validation", "AUC")

    # Save a check file so there is no ambiguity about the displayed legend values.
    check_df = pd.DataFrame([{
        "Best_SMOTE_Method": best_method,
        "Best_Sampling_Ratio": best_ratio,
        "PR_Legend_Train_mean_AUPRC": train_auprc_mean,
        "PR_Legend_Validation_mean_AUPRC": val_auprc_mean,
        "ROC_Legend_Train_mean_AUC": train_auc_mean,
        "ROC_Legend_Validation_mean_AUC": val_auc_mean,
        "Threshold": THRESHOLD,
        "Note": "Open PR_ROC_curves_FINAL/OPEN_THIS_PR_curve... The legend values are copied from 01_all_results.xlsx best row.",
    }])
    save_excel_formatted(check_df, os.path.join(output_dir, "00_OPEN_THIS_CHECK_VALUES.xlsx"), "Check")

    # PR curve
    train_precision, train_recall, _ = precision_recall_curve(best_obj["train_y"], best_obj["train_prob"])
    val_precision, val_recall, _ = precision_recall_curve(best_obj["val_y"], best_obj["val_prob"])

    fig, ax = plt.subplots(figsize=(6, 5), dpi=600)
    ax.plot(
        train_recall,
        train_precision,
        linewidth=1.8,
        color=COLOR_TRAIN,
        label=f"Train mean AUPRC = {train_auprc_mean:.3f}",
    )
    ax.plot(
        val_recall,
        val_precision,
        linewidth=1.8,
        color=COLOR_VAL,
        label=f"Validation mean AUPRC = {val_auprc_mean:.3f}",
    )
    ax.set_xlabel("Recall")
    ax.set_ylabel("Precision")
    ax.set_title(f"{MODEL_SHORT} Precision-Recall Curve")
    ax.legend(frameon=False, loc="best")
    sci_axes(ax)
    fig.tight_layout()
    pr_name = f"OPEN_THIS_PR_curve_BEST_{best_method}_ratio_{ratio_tag}_ValAUPRC_{val_auprc_mean:.3f}".replace(".", "p")
    save_figure(fig, fig_dir, pr_name)

    # ROC curve
    train_fpr, train_tpr, _ = roc_curve(best_obj["train_y"], best_obj["train_prob"])
    val_fpr, val_tpr, _ = roc_curve(best_obj["val_y"], best_obj["val_prob"])

    fig, ax = plt.subplots(figsize=(6, 5), dpi=600)
    ax.plot(
        train_fpr,
        train_tpr,
        linewidth=1.8,
        color=COLOR_TRAIN,
        label=f"Train mean AUC = {train_auc_mean:.3f}",
    )
    ax.plot(
        val_fpr,
        val_tpr,
        linewidth=1.8,
        color=COLOR_VAL,
        label=f"Validation mean AUC = {val_auc_mean:.3f}",
    )
    ax.plot([0, 1], [0, 1], linestyle="--", linewidth=1.0, color="gray")
    ax.set_xlabel("False Positive Rate")
    ax.set_ylabel("True Positive Rate")
    ax.set_title(f"{MODEL_SHORT} ROC Curve")
    ax.legend(frameon=False, loc="best")
    sci_axes(ax)
    fig.tight_layout()
    roc_name = f"OPEN_THIS_ROC_curve_BEST_{best_method}_ratio_{ratio_tag}_ValAUC_{val_auc_mean:.3f}".replace(".", "p")
    save_figure(fig, fig_dir, roc_name)


def bootstrap_fold_metrics(fold_df, output_dir, best_model_name=MODEL_SHORT):
    """Bootstrap 95% CI from 5 fold-level metric values after model selection."""
    rng = np.random.default_rng(BOOTSTRAP_RANDOM_STATE)
    metric_cols = [
        "AUPRC", "AUC", "Accuracy", "Balanced_Accuracy", "Recall_Sensitivity",
        "Specificity", "Precision", "F1", "MCC",
    ]

    records = []
    alpha = (100 - BOOTSTRAP_CI) / 2

    for dataset in ["Train", "Validation"]:
        sub = fold_df[fold_df["Dataset"] == dataset].reset_index(drop=True)
        n = len(sub)
        for metric in metric_cols:
            values = sub[metric].astype(float).values
            boot_means = []
            for _ in range(BOOTSTRAP_N):
                sample = rng.choice(values, size=n, replace=True)
                boot_means.append(np.nanmean(sample))
            lower, upper = np.nanpercentile(boot_means, [alpha, 100 - alpha])
            mean_value = float(np.nanmean(values))
            records.append({
                "Model": best_model_name,
                "Dataset": dataset,
                "Metric": metric,
                "Mean": mean_value,
                "SD": float(np.nanstd(values, ddof=1)) if n > 1 else 0.0,
                "CI_Lower": float(lower),
                "CI_Upper": float(upper),
                "Formatted_Percent_95CI": percent_ci_format(mean_value, lower, upper)
                if metric != "MCC" else f"{mean_value:.3f} ({lower:.3f}–{upper:.3f})",
                "N_Folds": n,
                "Bootstrap_N": BOOTSTRAP_N,
                "Threshold": THRESHOLD,
            })

    ci_df = pd.DataFrame(records)
    save_excel_formatted(ci_df, os.path.join(output_dir, "05_best_fold_bootstrap_95CI.xlsx"), "Bootstrap_95CI")

    val_ci = ci_df[ci_df["Dataset"] == "Validation"].set_index("Metric")
    paper_row = pd.DataFrame([{
        "Model": best_model_name,
        "Balanced accuracy": val_ci.loc["Balanced_Accuracy", "Formatted_Percent_95CI"],
        "Recall": val_ci.loc["Recall_Sensitivity", "Formatted_Percent_95CI"],
        "Precision": val_ci.loc["Precision", "Formatted_Percent_95CI"],
        "F1 score": val_ci.loc["F1", "Formatted_Percent_95CI"],
        "AUPRC": val_ci.loc["AUPRC", "Formatted_Percent_95CI"],
        "AUROC": val_ci.loc["AUC", "Formatted_Percent_95CI"],
        "Threshold": THRESHOLD,
        "Note": f"Values are bootstrap {BOOTSTRAP_CI}% confidence intervals based on five fold-level cross-validation metrics.",
    }])
    save_excel_formatted(paper_row, os.path.join(output_dir, "06_paper_table_row_validation_95CI.xlsx"), "Paper_Table")

    return ci_df, paper_row


# =============================================================================
# 4. SHAP for best CatBoost model
# =============================================================================
def run_catboost_shap(X, y, best_sampler_name, best_sampler_class, best_ratio, output_dir):
    shap_dir = os.path.join(output_dir, "SHAP_CatBoost")
    ensure_output_dir(shap_dir)

    sampler = make_sampler(best_sampler_name, best_sampler_class, best_ratio, y)
    X_res, y_res = sampler.fit_resample(X, y)

    model = build_model()
    model.fit(X_res, y_res)

    try:
        import shap
    except Exception as e:
        err_df = pd.DataFrame({"Error": [f"shap未安装或导入失败: {e}"]})
        save_excel_formatted(err_df, os.path.join(shap_dir, "SHAP_error.xlsx"), "Error")

        imp_values = model.get_feature_importance()
        imp_df = pd.DataFrame({
            "Feature": X.columns,
            "CatBoost_Importance": imp_values,
        }).sort_values("CatBoost_Importance", ascending=False).reset_index(drop=True)
        save_excel_formatted(imp_df, os.path.join(shap_dir, "CatBoost_feature_importance_fallback.xlsx"), "Importance")

        plot_df = imp_df.head(TOP_N_FEATURES).sort_values("CatBoost_Importance", ascending=True)
        height = max(4.5, 0.38 * len(plot_df))
        fig, ax = plt.subplots(figsize=(8, height), dpi=600)
        ax.barh(plot_df["Feature"], plot_df["CatBoost_Importance"], color=COLOR_BAR, edgecolor="black", linewidth=0.4)
        ax.set_xlabel("CatBoost feature importance")
        ax.set_ylabel("")
        ax.set_title(f"Top {min(TOP_N_FEATURES, len(plot_df))} CatBoost Features")
        sci_axes(ax)
        fig.tight_layout()
        save_figure(fig, shap_dir, "CatBoost_feature_importance_top10_fallback")
        return imp_df

    explainer = shap.TreeExplainer(model)
    shap_values = explainer.shap_values(X)

    if isinstance(shap_values, list):
        shap_values_to_use = shap_values[1]
    else:
        shap_values_to_use = shap_values

    shap_df = pd.DataFrame(shap_values_to_use, columns=X.columns)
    save_excel_formatted(shap_df, os.path.join(shap_dir, "SHAP_values.xlsx"), "SHAP_Values")

    mean_abs = np.abs(shap_values_to_use).mean(axis=0)
    imp_df = pd.DataFrame({
        "Feature": X.columns,
        "Mean_abs_SHAP": mean_abs,
    }).sort_values("Mean_abs_SHAP", ascending=False).reset_index(drop=True)
    save_excel_formatted(imp_df, os.path.join(shap_dir, "SHAP_feature_importance.xlsx"), "SHAP_Importance")

    # SCI-style SHAP top10 horizontal bar
    plot_df = imp_df.head(TOP_N_FEATURES).sort_values("Mean_abs_SHAP", ascending=True)
    height = max(4.5, 0.38 * len(plot_df))
    fig, ax = plt.subplots(figsize=(8, height), dpi=600)
    ax.barh(plot_df["Feature"], plot_df["Mean_abs_SHAP"], color=COLOR_BAR, edgecolor="black", linewidth=0.4)
    ax.set_xlabel("Mean absolute SHAP value")
    ax.set_ylabel("")
    ax.set_title(f"Top {min(TOP_N_FEATURES, len(plot_df))} CatBoost SHAP Features")
    sci_axes(ax)
    fig.tight_layout()
    save_figure(fig, shap_dir, "SHAP_top10_feature_importance")

    # SHAP beeswarm
    plt.figure(figsize=(7, 5), dpi=600)
    shap.summary_plot(shap_values_to_use, X, show=False, max_display=min(20, X.shape[1]))
    plt.tight_layout()
    plt.savefig(os.path.join(shap_dir, "SHAP_beeswarm.png"), dpi=600, bbox_inches="tight")
    plt.savefig(os.path.join(shap_dir, "SHAP_beeswarm.pdf"), bbox_inches="tight")
    plt.close()

    return imp_df


# =============================================================================
# 5. Main
# =============================================================================
def main():
    OUTPUT_DIR = make_timestamped_output_dir()
    X, y = load_data(DATA_FILE, TARGET_COL)

    print("=" * 80)
    print("RUNNING CLEAN CATBOOST FINAL SCRIPT")
    print(f"Output folder: {OUTPUT_DIR}")
    print(f"N={len(y)}, Positive={int(y.sum())}, Negative={int(len(y) - y.sum())}")
    print(f"Positive ratio={y.mean():.4f}")
    print(f"Model params={MODEL_PARAMS}")
    print(f"SMOTE k_neighbors={SMOTE_K_NEIGHBORS}")
    print(f"Threshold={THRESHOLD}")
    print("=" * 80)

    all_results = []
    all_folds = []
    summaries = {}
    full_objs = {}

    current_ratio = y.sum() / (len(y) - y.sum())

    for sampler_name, sampler_class in SAMPLER_CONFIGS.items():
        for ratio in SAMPLING_RATIOS:
            if ratio <= current_ratio:
                print(f"Skip {sampler_name} ratio={ratio}: ratio <= current minority/majority ratio")
                continue

            print(f"Running: {sampler_name}, ratio={ratio}")
            obj = run_one_setting(X, y, sampler_name, sampler_class, ratio)

            all_results.append({
                "Model": obj["Model"],
                "SMOTE_Method": obj["SMOTE_Method"],
                "Sampling_Ratio": obj["Sampling_Ratio"],
                "Train_AUPRC": obj["Train_AUPRC"],
                "Val_AUPRC": obj["Val_AUPRC"],
                "Gap": obj["Gap"],
                "Threshold": obj["Threshold"],
                "Params": obj["Params"],
            })
            all_folds.append(obj["Fold_DF"])
            key = f"{sampler_name}_{ratio}"
            summaries[key] = obj["Summary_DF"]
            full_objs[key] = obj

    if len(all_results) == 0:
        raise RuntimeError("没有任何SMOTE设置被运行。请检查 SAMPLING_RATIOS 是否大于当前少数类/多数类比例。")

    result_df = pd.DataFrame(all_results)
    result_df = add_selection_flags(result_df)
    result_df = result_df.sort_values(["Val_AUPRC", "Gap"], ascending=[False, True]).reset_index(drop=True)

    best = result_df.iloc[0]
    best_sampler_name = best["SMOTE_Method"]
    best_ratio = float(best["Sampling_Ratio"])
    best_key = f"{best_sampler_name}_{best_ratio}"
    best_sampler_class = SAMPLER_CONFIGS[best_sampler_name]
    best_obj = full_objs[best_key]
    best_summary = summaries[best_key]

    # Safety checks: these must be identical.
    best_summary_val_auprc = get_mean_metric(best_summary, "Validation", "AUPRC")
    if abs(float(best["Val_AUPRC"]) - best_summary_val_auprc) > 1e-12:
        raise RuntimeError(
            f"Internal mismatch: best Val_AUPRC={best['Val_AUPRC']} but best_summary Val_AUPRC={best_summary_val_auprc}"
        )

    all_fold_df = pd.concat(all_folds, ignore_index=True)

    # Core Excel outputs
    save_excel_formatted(result_df, os.path.join(OUTPUT_DIR, "01_all_results.xlsx"), "All_Results")
    save_excel_formatted(all_fold_df, os.path.join(OUTPUT_DIR, "02_fold_metrics_all_settings.xlsx"), "Fold_Metrics_All")
    save_excel_formatted(best_obj["Fold_DF"], os.path.join(OUTPUT_DIR, "02_best_fold_metrics_ONLY.xlsx"), "Best_Folds")
    save_excel_formatted(best_summary, os.path.join(OUTPUT_DIR, "03_best_summary_metrics_mean_sd.xlsx"), "Summary")
    save_excel_formatted(result_df.head(1), os.path.join(OUTPUT_DIR, "04_best_setting.xlsx"), "Best_Setting")

    # Combined PR/ROC curves for selected best setting only.
    plot_combined_pr_roc(best_obj, best, best_summary, OUTPUT_DIR)

    # Bootstrap 95% CI after best setting is selected.
    ci_df, paper_row = bootstrap_fold_metrics(best_obj["Fold_DF"], OUTPUT_DIR, MODEL_SHORT)

    # SHAP interpretation for selected best setting only.
    run_catboost_shap(X, y, best_sampler_name, best_sampler_class, best_ratio, OUTPUT_DIR)

    print("=" * 80)
    print("FINISHED.")
    print(f"IMPORTANT: Open this NEW output folder only:\n{OUTPUT_DIR}")
    print("Best setting:")
    print(result_df.head(1).to_string(index=False))
    print("Best summary:")
    print(best_summary.to_string(index=False))
    print("Paper table row:")
    print(paper_row.to_string(index=False))
    print("=" * 80)


if __name__ == "__main__":
    main()
