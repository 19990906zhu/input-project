
import os
import ntpath
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import (
    average_precision_score,
    precision_recall_curve,
    roc_auc_score,
    roc_curve,
    accuracy_score,
    balanced_accuracy_score,
    recall_score,
    precision_score,
    f1_score,
    confusion_matrix,
    matthews_corrcoef,
)
from sklearn.feature_selection import RFECV
from imblearn.ensemble import BalancedRandomForestClassifier

warnings.filterwarnings("ignore")


# =============================================================================
# 0. Basic settings: mainly edit here
# =============================================================================
DATA_FILE = r"D:\萝卜投稿专用\Original Research - Spinal Infection Following Internal Fixation Surgery\插值选择\Xgboost\lr=0.1\训练集-z-scores.xlsx"
TARGET_COL = "Infection"
POS_LABEL = 1

RANDOM_STATE = 42
N_SPLITS = 5
THRESHOLD = 0.30
TOP_N_FEATURES = 10

MODEL_NAME = "Balanced Random Forest"
MODEL_SHORT = "BalancedRF"

# Output base folder. A timestamped subfolder will be created automatically.
OUTPUT_BASE_DIR = r"D:\萝卜投稿专用\Original Research - Spinal Infection Following Internal Fixation Surgery\插值选择\Xgboost\lr=0.1\建模-1"

MODEL_PARAMS = {
    "n_estimators": 48,
    "max_depth": 1,
    "min_samples_leaf": 4,
    "replacement": True,
    "sampling_strategy": "all",
    "random_state": RANDOM_STATE,
    "n_jobs": -1,
}

BOOTSTRAP_N = 2000
BOOTSTRAP_CI = 95
BOOTSTRAP_RANDOM_STATE = 2026

# =============================================================================
# 1. Plot style
# =============================================================================
FIG_DPI = 600
COLOR_TRAIN = "#4E79A7"
COLOR_VAL = "#E15759"
COLOR_BAR = "#4E79A7"
COLOR_GREY = "#808080"

plt.rcParams.update({
    "font.family": ["Arial", "DejaVu Sans"],
    "font.size": 10,
    "axes.labelsize": 11,
    "axes.titlesize": 13,
    "legend.fontsize": 10,
    "xtick.labelsize": 10,
    "ytick.labelsize": 10,
    "axes.linewidth": 0.8,
    "pdf.fonttype": 42,
    "ps.fonttype": 42,
})


# =============================================================================
# 2. Utilities
# =============================================================================
def ensure_output_dir(path):
    if path:
        os.makedirs(path, exist_ok=True)


def make_timestamped_output_dir():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = os.path.join(OUTPUT_BASE_DIR, f"BalancedRF_FINAL_BOOTSTRAP_RFECV_{timestamp}")
    ensure_output_dir(out_dir)
    return out_dir


def resolve_data_file(file_path):
    """Try original path first, then current folder and script folder by basename."""
    if os.path.exists(file_path):
        return file_path

    base_names = []
    base_names.append(os.path.basename(file_path))
    base_names.append(ntpath.basename(file_path))
    base_names = [b for b in dict.fromkeys(base_names) if b]

    search_dirs = [os.getcwd()]
    try:
        search_dirs.append(os.path.dirname(os.path.abspath(__file__)))
    except Exception:
        pass

    for d in search_dirs:
        for b in base_names:
            candidate = os.path.join(d, b)
            if os.path.exists(candidate):
                return candidate

    raise FileNotFoundError(
        f"找不到数据文件：{file_path}\n"
        f"请确认Excel文件与本py在同一文件夹，或把 DATA_FILE 改成真实路径。"
    )


def load_data(file_path, target_col):
    file_path = resolve_data_file(file_path)
    df = pd.read_excel(file_path)

    if target_col not in df.columns:
        raise ValueError(f"找不到结局变量列：{target_col}；当前列名：{list(df.columns)}")

    df = df.dropna(axis=1, how="all")
    df = df.dropna(subset=[target_col]).copy()

    y = pd.to_numeric(df[target_col], errors="coerce")
    X = df.drop(columns=[target_col])

    non_numeric_cols = X.select_dtypes(exclude=[np.number]).columns.tolist()
    if non_numeric_cols:
        print("以下非数值列将被自动删除：")
        print(non_numeric_cols)
        X = X.drop(columns=non_numeric_cols)

    if X.isna().sum().sum() > 0:
        print("警告：X中存在缺失值，将用列中位数填补。")
        X = X.fillna(X.median(numeric_only=True))

    valid = y.notna()
    X = X.loc[valid].reset_index(drop=True)
    y = y.loc[valid].astype(int).reset_index(drop=True)
    y = (y == POS_LABEL).astype(int)

    if y.nunique() != 2:
        raise ValueError("结局变量必须包含0和1两类。请检查 Infection 列。")

    return X, y, file_path


def build_model():
    return BalancedRandomForestClassifier(**MODEL_PARAMS)


def get_positive_probability(model, X_data):
    if hasattr(model, "predict_proba"):
        return model.predict_proba(X_data)[:, 1]
    if hasattr(model, "decision_function"):
        score = model.decision_function(X_data)
        return 1 / (1 + np.exp(-score))
    return model.predict(X_data)


def calc_metrics(y_true, y_prob, threshold=THRESHOLD):
    y_pred = (y_prob >= threshold).astype(int)
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    specificity = tn / (tn + fp) if (tn + fp) > 0 else np.nan

    return {
        "AUPRC": average_precision_score(y_true, y_prob),
        "AUC": roc_auc_score(y_true, y_prob),
        "Accuracy": accuracy_score(y_true, y_pred),
        "Balanced_Accuracy": balanced_accuracy_score(y_true, y_pred),
        "Recall_Sensitivity": recall_score(y_true, y_pred, zero_division=0),
        "Specificity": specificity,
        "Precision": precision_score(y_true, y_pred, zero_division=0),
        "F1": f1_score(y_true, y_pred, zero_division=0),
        "MCC": matthews_corrcoef(y_true, y_pred),
        "TP": tp,
        "FP": fp,
        "TN": tn,
        "FN": fn,
    }


def summarize_fold_metrics(fold_df):
    ignore_cols = ["Fold", "Dataset"]
    metrics = [c for c in fold_df.columns if c not in ignore_cols]
    rows = []
    for dataset in ["Train", "Validation"]:
        sub = fold_df[fold_df["Dataset"] == dataset]
        row = {"Dataset": dataset}
        for m in metrics:
            row[m + "_Mean"] = float(sub[m].mean())
            row[m + "_SD"] = float(sub[m].std(ddof=1)) if len(sub) > 1 else 0.0
        rows.append(row)
    return pd.DataFrame(rows)


def get_mean_metric(summary_df, dataset, metric):
    return float(summary_df.loc[summary_df["Dataset"] == dataset, f"{metric}_Mean"].iloc[0])


def save_excel_formatted(df, path, sheet_name="Sheet1"):
    ensure_output_dir(os.path.dirname(path))
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        ws = writer.sheets[sheet_name[:31]]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="center")
                cell.border = border
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


def sci_axes(ax):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(axis="both", direction="out", length=3, width=0.8)
    ax.grid(False)


def save_figure(fig, out_dir, name):
    ensure_output_dir(out_dir)
    fig.savefig(os.path.join(out_dir, f"{name}.png"), dpi=FIG_DPI, bbox_inches="tight")
    fig.savefig(os.path.join(out_dir, f"{name}.pdf"), bbox_inches="tight")
    plt.close(fig)


def percent_ci_format(mean, lower, upper):
    return f"{100 * mean:.2f}% ({100 * lower:.2f}–{100 * upper:.2f}%)"


# =============================================================================
# 3. CV, plots, bootstrap
# =============================================================================
def run_cv(X, y):
    cv = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)
    fold_rows = []
    val_y_all, val_prob_all = [], []
    train_y_all, train_prob_all = [], []

    for fold_id, (train_idx, val_idx) in enumerate(cv.split(X, y), start=1):
        X_train, X_val = X.iloc[train_idx], X.iloc[val_idx]
        y_train, y_val = y.iloc[train_idx], y.iloc[val_idx]

        model = build_model()
        model.fit(X_train, y_train)

        train_prob = get_positive_probability(model, X_train)
        val_prob = get_positive_probability(model, X_val)

        for dataset, yy, pp in [
            ("Train", y_train, train_prob),
            ("Validation", y_val, val_prob),
        ]:
            row = {"Fold": fold_id, "Dataset": dataset}
            row.update(calc_metrics(yy, pp, THRESHOLD))
            fold_rows.append(row)

        train_y_all.extend(y_train.tolist())
        train_prob_all.extend(train_prob.tolist())
        val_y_all.extend(y_val.tolist())
        val_prob_all.extend(val_prob.tolist())

    fold_df = pd.DataFrame(fold_rows)
    summary_df = summarize_fold_metrics(fold_df)

    return {
        "Fold_DF": fold_df,
        "Summary_DF": summary_df,
        "train_y": np.array(train_y_all),
        "train_prob": np.array(train_prob_all),
        "val_y": np.array(val_y_all),
        "val_prob": np.array(val_prob_all),
    }


def plot_combined_pr_roc(cv_obj, output_dir):
    fig_dir = os.path.join(output_dir, "PR_ROC_curves_FINAL")
    ensure_output_dir(fig_dir)

    summary_df = cv_obj["Summary_DF"]
    train_y, train_prob = cv_obj["train_y"], cv_obj["train_prob"]
    val_y, val_prob = cv_obj["val_y"], cv_obj["val_prob"]

    train_auprc_mean = get_mean_metric(summary_df, "Train", "AUPRC")
    val_auprc_mean = get_mean_metric(summary_df, "Validation", "AUPRC")
    train_auc_mean = get_mean_metric(summary_df, "Train", "AUC")
    val_auc_mean = get_mean_metric(summary_df, "Validation", "AUC")

    tag = f"ValAUPRC_{val_auprc_mean:.3f}".replace(".", "p")

    # PR curve: curve uses pooled CV predictions; legend uses 5-fold mean to match Excel.
    train_precision, train_recall, _ = precision_recall_curve(train_y, train_prob)
    val_precision, val_recall, _ = precision_recall_curve(val_y, val_prob)

    fig, ax = plt.subplots(figsize=(6, 5), dpi=FIG_DPI)
    ax.plot(train_recall, train_precision, linewidth=1.8, color=COLOR_TRAIN,
            label=f"Train mean AUPRC = {train_auprc_mean:.3f}")
    ax.plot(val_recall, val_precision, linewidth=1.8, color=COLOR_VAL,
            label=f"Validation mean AUPRC = {val_auprc_mean:.3f}")
    ax.set_xlabel("Recall")
    ax.set_ylabel("Precision")
    ax.set_title(f"{MODEL_NAME} Precision-Recall Curve")
    ax.legend(frameon=False, loc="best")
    ax.set_xlim(-0.02, 1.02)
    ax.set_ylim(-0.02, 1.02)
    sci_axes(ax)
    fig.tight_layout()
    save_figure(fig, fig_dir, f"OPEN_THIS_PR_curve_{MODEL_SHORT}_{tag}")

    # ROC curve
    train_fpr, train_tpr, _ = roc_curve(train_y, train_prob)
    val_fpr, val_tpr, _ = roc_curve(val_y, val_prob)

    fig, ax = plt.subplots(figsize=(6, 5), dpi=FIG_DPI)
    ax.plot(train_fpr, train_tpr, linewidth=1.8, color=COLOR_TRAIN,
            label=f"Train mean AUC = {train_auc_mean:.3f}")
    ax.plot(val_fpr, val_tpr, linewidth=1.8, color=COLOR_VAL,
            label=f"Validation mean AUC = {val_auc_mean:.3f}")
    ax.plot([0, 1], [0, 1], linestyle="--", linewidth=1.0, color=COLOR_GREY)
    ax.set_xlabel("False Positive Rate")
    ax.set_ylabel("True Positive Rate")
    ax.set_title(f"{MODEL_NAME} ROC Curve")
    ax.legend(frameon=False, loc="lower right")
    ax.set_xlim(-0.02, 1.02)
    ax.set_ylim(-0.02, 1.02)
    sci_axes(ax)
    fig.tight_layout()
    save_figure(fig, fig_dir, f"OPEN_THIS_ROC_curve_{MODEL_SHORT}_{tag}")

    check_df = pd.DataFrame([{
        "Model": MODEL_NAME,
        "Threshold": THRESHOLD,
        "Train_AUPRC_mean_for_PR_legend": train_auprc_mean,
        "Validation_AUPRC_mean_for_PR_legend": val_auprc_mean,
        "Train_AUC_mean_for_ROC_legend": train_auc_mean,
        "Validation_AUC_mean_for_ROC_legend": val_auc_mean,
        "Open_PR_Figure_Name_Contains": f"OPEN_THIS_PR_curve_{MODEL_SHORT}_{tag}",
        "Open_ROC_Figure_Name_Contains": f"OPEN_THIS_ROC_curve_{MODEL_SHORT}_{tag}",
    }])
    save_excel_formatted(check_df, os.path.join(output_dir, "00_OPEN_THIS_CHECK_VALUES.xlsx"), "Check_Values")


def bootstrap_fold_metrics(fold_df, output_dir):
    rng = np.random.default_rng(BOOTSTRAP_RANDOM_STATE)
    metric_cols = [
        "AUPRC", "AUC", "Accuracy", "Balanced_Accuracy", "Recall_Sensitivity",
        "Specificity", "Precision", "F1", "MCC",
    ]

    records = []
    alpha = (100 - BOOTSTRAP_CI) / 2

    for dataset in ["Train", "Validation"]:
        sub = fold_df[fold_df["Dataset"] == dataset].reset_index(drop=True)
        n = len(sub)
        for metric in metric_cols:
            values = sub[metric].astype(float).values
            boot_means = []
            for _ in range(BOOTSTRAP_N):
                sample = rng.choice(values, size=n, replace=True)
                boot_means.append(np.nanmean(sample))
            lower, upper = np.nanpercentile(boot_means, [alpha, 100 - alpha])
            mean_value = float(np.nanmean(values))
            records.append({
                "Model": MODEL_NAME,
                "Dataset": dataset,
                "Metric": metric,
                "Mean": mean_value,
                "SD": float(np.nanstd(values, ddof=1)) if n > 1 else 0.0,
                "CI_Lower": float(lower),
                "CI_Upper": float(upper),
                "Formatted_Percent_95CI": percent_ci_format(mean_value, lower, upper)
                if metric != "MCC" else f"{mean_value:.3f} ({lower:.3f}–{upper:.3f})",
                "N_Folds": n,
                "Bootstrap_N": BOOTSTRAP_N,
                "Threshold": THRESHOLD,
            })

    ci_df = pd.DataFrame(records)
    save_excel_formatted(ci_df, os.path.join(output_dir, "03_bootstrap_95CI.xlsx"), "Bootstrap_95CI")

    val_ci = ci_df[ci_df["Dataset"] == "Validation"].set_index("Metric")
    paper_row = pd.DataFrame([{
        "Model": MODEL_NAME,
        "Balanced accuracy": val_ci.loc["Balanced_Accuracy", "Formatted_Percent_95CI"],
        "Recall": val_ci.loc["Recall_Sensitivity", "Formatted_Percent_95CI"],
        "Precision": val_ci.loc["Precision", "Formatted_Percent_95CI"],
        "F1 score": val_ci.loc["F1", "Formatted_Percent_95CI"],
        "AUPRC": val_ci.loc["AUPRC", "Formatted_Percent_95CI"],
        "AUROC": val_ci.loc["AUC", "Formatted_Percent_95CI"],
        "Threshold": THRESHOLD,
        "Note": f"Values are bootstrap {BOOTSTRAP_CI}% confidence intervals based on five-fold cross-validation metrics.",
    }])
    save_excel_formatted(paper_row, os.path.join(output_dir, "04_paper_table_row_validation_95CI.xlsx"), "Paper_Table")

    return ci_df, paper_row


# =============================================================================
# 4. RFECV-selected feature importance
# =============================================================================
def run_rfecv_and_plot_top10(X, y, output_dir):
    rfecv_dir = os.path.join(output_dir, "RFECV")
    ensure_output_dir(rfecv_dir)

    estimator = build_model()
    cv = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)

    rfecv = RFECV(
        estimator=estimator,
        step=1,
        cv=cv,
        scoring="average_precision",
        n_jobs=-1,
        importance_getter="feature_importances_",
        min_features_to_select=1,
    )
    rfecv.fit(X, y)

    ranking = pd.DataFrame({
        "Feature": X.columns,
        "Selected_by_RFECV": rfecv.support_,
        "RFECV_Ranking": rfecv.ranking_,
    }).sort_values(["RFECV_Ranking", "Feature"], ascending=[True, True]).reset_index(drop=True)

    selected_features = X.columns[rfecv.support_].tolist()
    selected_df = pd.DataFrame({"Selected_Feature": selected_features})

    save_excel_formatted(selected_df, os.path.join(rfecv_dir, "RFECV_selected_features.xlsx"), "Selected")
    save_excel_formatted(ranking, os.path.join(rfecv_dir, "RFECV_feature_ranking.xlsx"), "RFECV_Ranking")

    cv_scores = pd.DataFrame({
        "Number_of_Selected_Features": np.arange(1, len(rfecv.cv_results_["mean_test_score"]) + 1),
        "Mean_CV_AUPRC": rfecv.cv_results_["mean_test_score"],
        "SD_CV_AUPRC": rfecv.cv_results_.get("std_test_score", np.nan),
    })
    save_excel_formatted(cv_scores, os.path.join(rfecv_dir, "RFECV_cv_scores.xlsx"), "CV_Scores")

    # Correct display: selected features first, then refit model on selected features and plot importance.
    X_selected = X[selected_features]
    final_model = build_model()
    final_model.fit(X_selected, y)

    rfecv_imp = pd.DataFrame({
        "Feature": selected_features,
        "Importance_after_RFECV": final_model.feature_importances_,
    }).sort_values("Importance_after_RFECV", ascending=False).reset_index(drop=True)

    save_excel_formatted(
        rfecv_imp,
        os.path.join(rfecv_dir, "RFECV_selected_feature_importance.xlsx"),
        "RFECV_Importance",
    )

    plot_df = rfecv_imp.head(TOP_N_FEATURES).sort_values("Importance_after_RFECV", ascending=True)
    height = max(4.5, 0.38 * len(plot_df))
    fig, ax = plt.subplots(figsize=(8, height), dpi=FIG_DPI)
    ax.barh(
        plot_df["Feature"],
        plot_df["Importance_after_RFECV"],
        color=COLOR_BAR,
        edgecolor="black",
        linewidth=0.4,
    )
    ax.set_xlabel("Feature importance after RFECV")
    ax.set_ylabel("")
    ax.set_title(f"Top {min(TOP_N_FEATURES, len(plot_df))} RFECV-Selected Features")
    sci_axes(ax)
    fig.tight_layout()
    save_figure(fig, rfecv_dir, "OPEN_THIS_RFECV_top10_selected_features")

    return ranking, rfecv_imp


# =============================================================================
# 5. Main
# =============================================================================
def main():
    output_dir = make_timestamped_output_dir()
    X, y, resolved_file = load_data(DATA_FILE, TARGET_COL)

    print("=" * 80)
    print(f"Model: {MODEL_NAME}")
    print(f"Resolved data file: {resolved_file}")
    print(f"N={len(y)}, Positive={int(y.sum())}, Negative={int(len(y) - y.sum())}")
    print(f"Positive ratio={y.mean():.4f}")
    print(f"Threshold={THRESHOLD}")
    print(f"Model params={MODEL_PARAMS}")
    print(f"Output folder: {output_dir}")
    print("=" * 80)

    cv_obj = run_cv(X, y)
    fold_df = cv_obj["Fold_DF"]
    summary_df = cv_obj["Summary_DF"]

    save_excel_formatted(fold_df, os.path.join(output_dir, "01_fold_metrics.xlsx"), "Fold_Metrics")
    save_excel_formatted(summary_df, os.path.join(output_dir, "02_summary_metrics_mean_sd.xlsx"), "Summary")

    plot_combined_pr_roc(cv_obj, output_dir)
    bootstrap_fold_metrics(fold_df, output_dir)
    run_rfecv_and_plot_top10(X, y, output_dir)

    print("=" * 80)
    print("Finished.")
    print("IMPORTANT: Open this NEW output folder only:")
    print(output_dir)
    print("Key files:")
    print("  00_OPEN_THIS_CHECK_VALUES.xlsx")
    print("  01_fold_metrics.xlsx")
    print("  02_summary_metrics_mean_sd.xlsx")
    print("  04_paper_table_row_validation_95CI.xlsx")
    print("  PR_ROC_curves_FINAL/OPEN_THIS_PR_curve_*.png")
    print("  RFECV/OPEN_THIS_RFECV_top10_selected_features.png")
    print("=" * 80)
    print(summary_df.to_string(index=False))


if __name__ == "__main__":
    main()
