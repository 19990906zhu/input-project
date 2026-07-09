
import os
import ntpath
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from sklearn.base import clone
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import (
    average_precision_score,
    precision_recall_curve,
    roc_auc_score,
    roc_curve,
    accuracy_score,
    balanced_accuracy_score,
    recall_score,
    precision_score,
    f1_score,
    confusion_matrix,
    matthews_corrcoef,
)
from sklearn.feature_selection import RFECV
from sklearn.preprocessing import StandardScaler
from sklearn.svm import SVC

from imblearn.pipeline import Pipeline as ImbPipeline
from imblearn.over_sampling import SMOTE, BorderlineSMOTE, SVMSMOTE

warnings.filterwarnings("ignore")

# =============================================================================
# 0. Basic settings: mainly edit here
# =============================================================================
DATA_FILE = r"D:\萝卜投稿专用\Original Research - Spinal Infection Following Internal Fixation Surgery\插值选择\Xgboost\lr=0.1\P小于0.05_按临床动态标准共线性删除后.xlsx"

TARGET_COL = "Infection"
POS_LABEL = 1

RANDOM_STATE = 42
N_SPLITS = 5
THRESHOLD = 0.30
TOP_N_FEATURES = 10

MODEL_SHORT = "SVM"
USE_SCALER = True   # FIXED: scaler is fitted within each CV training fold; use raw/unstandardized input data.
SMOTE_K_NEIGHBORS = 20

# Original SVM ratios
SAMPLING_RATIOS = [0.61]

MODEL_PARAMS = {
    "C": 5,
    "gamma": 0.05,
    "kernel": "linear",
    "class_weight": 'balanced',
    "probability": True,
    "random_state": RANDOM_STATE,
}

SAMPLER_CONFIGS = {
    "SMOTE": SMOTE,
    "BorderlineSMOTE": BorderlineSMOTE,
    "SVMSMOTE": SVMSMOTE,
}

BOOTSTRAP_N = 2000
BOOTSTRAP_CI = 95
BOOTSTRAP_RANDOM_STATE = 2026

# =============================================================================
# 1. Output folder and plot style
# =============================================================================
RUN_TAG = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_DIR = r"D:\萝卜投稿专用\Original Research - Spinal Infection Following Internal Fixation Surgery\插值选择\Xgboost\lr=0.16\建模-1"

plt.rcParams["font.family"] = ["Arial", "DejaVu Sans"]
plt.rcParams["font.size"] = 10
plt.rcParams["axes.linewidth"] = 0.8
plt.rcParams["pdf.fonttype"] = 42
plt.rcParams["ps.fonttype"] = 42

COLOR_TRAIN = "#4E79A7"
COLOR_VAL = "#E15759"
COLOR_BAR = "#4E79A7"
COLOR_GREY = "#808080"


# =============================================================================
# 2. Utilities
# =============================================================================
def ensure_output_dir(path):
    if path:
        os.makedirs(path, exist_ok=True)


def resolve_data_file(file_path):
    """Try original path first, then current folder and script folder by basename."""
    if os.path.exists(file_path):
        return file_path

    base_names = []
    base_names.append(os.path.basename(file_path))
    base_names.append(ntpath.basename(file_path))
    base_names = [b for b in dict.fromkeys(base_names) if b]

    search_dirs = [os.getcwd()]
    try:
        search_dirs.append(os.path.dirname(os.path.abspath(__file__)))
    except Exception:
        pass

    for d in search_dirs:
        for b in base_names:
            candidate = os.path.join(d, b)
            if os.path.exists(candidate):
                return candidate

    raise FileNotFoundError(
        f"找不到数据文件：{file_path}\n"
        f"请确认Excel文件与本py在同一文件夹，或把 DATA_FILE 改成真实路径。"
    )


def load_data(file_path, target_col):
    file_path = resolve_data_file(file_path)
    df = pd.read_excel(file_path)

    if target_col not in df.columns:
        raise ValueError(f"找不到结局变量列：{target_col}；当前列名：{list(df.columns)}")

    df = df.dropna(axis=1, how="all")
    df = df.dropna(subset=[target_col]).copy()

    y = pd.to_numeric(df[target_col], errors="coerce")
    X = df.drop(columns=[target_col])

    non_numeric_cols = X.select_dtypes(exclude=[np.number]).columns.tolist()
    if non_numeric_cols:
        print("以下非数值列将被自动删除：")
        print(non_numeric_cols)
        X = X.drop(columns=non_numeric_cols)

    if X.isna().sum().sum() > 0:
        print("警告：X中存在缺失值，将用列中位数填补。")
        X = X.fillna(X.median(numeric_only=True))

    valid = y.notna()
    X = X.loc[valid].reset_index(drop=True)
    y = y.loc[valid].astype(int).reset_index(drop=True)
    y = (y == POS_LABEL).astype(int)

    if y.nunique() != 2:
        raise ValueError("结局变量必须包含0和1两类。请检查 Infection 列。")

    return X, y


def build_model():
    return SVC(**MODEL_PARAMS)


def safe_k_neighbors(y, requested_k=SMOTE_K_NEIGHBORS):
    minority_n = int(min(np.sum(y == 0), np.sum(y == 1)))
    return max(1, min(requested_k, minority_n - 1))


def make_sampler(sampler_name, sampler_class, ratio, y=None):
    k = SMOTE_K_NEIGHBORS if y is None else safe_k_neighbors(np.asarray(y), requested_k=SMOTE_K_NEIGHBORS)
    kwargs = {
        "sampling_strategy": ratio,
        "random_state": RANDOM_STATE,
        "k_neighbors": k,
    }
    if sampler_name == "BorderlineSMOTE":
        return sampler_class(**kwargs, kind="borderline-1")
    if sampler_name == "SVMSMOTE":
        return sampler_class(**kwargs, m_neighbors=min(10, max(1, k)))
    return sampler_class(**kwargs)


def make_pipeline(sampler, model):
    steps = []
    if USE_SCALER:
        steps.append(("scaler", StandardScaler()))
    steps.append(("smote", sampler))
    steps.append(("model", model))
    return ImbPipeline(steps=steps)


def get_positive_probability(model, X_data):
    if hasattr(model, "predict_proba"):
        return model.predict_proba(X_data)[:, 1]
    if hasattr(model, "decision_function"):
        score = model.decision_function(X_data)
        return 1 / (1 + np.exp(-score))
    return model.predict(X_data)


def calc_metrics(y_true, y_prob, threshold=THRESHOLD):
    y_pred = (y_prob >= threshold).astype(int)
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    specificity = tn / (tn + fp) if (tn + fp) > 0 else np.nan

    return {
        "AUPRC": average_precision_score(y_true, y_prob),
        "AUC": roc_auc_score(y_true, y_prob),
        "Accuracy": accuracy_score(y_true, y_pred),
        "Balanced_Accuracy": balanced_accuracy_score(y_true, y_pred),
        "Recall_Sensitivity": recall_score(y_true, y_pred, zero_division=0),
        "Specificity": specificity,
        "Precision": precision_score(y_true, y_pred, zero_division=0),
        "F1": f1_score(y_true, y_pred, zero_division=0),
        "MCC": matthews_corrcoef(y_true, y_pred),
        "TP": tp,
        "FP": fp,
        "TN": tn,
        "FN": fn,
    }


def summarize_fold_metrics(fold_df):
    ignore_cols = ["Fold", "Dataset", "SMOTE_Method", "Sampling_Ratio"]
    metrics = [c for c in fold_df.columns if c not in ignore_cols]
    rows = []
    for dataset in ["Train", "Validation"]:
        sub = fold_df[fold_df["Dataset"] == dataset]
        row = {"Dataset": dataset}
        for m in metrics:
            row[m + "_Mean"] = float(sub[m].mean())
            row[m + "_SD"] = float(sub[m].std(ddof=1)) if len(sub) > 1 else 0.0
        rows.append(row)
    return pd.DataFrame(rows)


def get_mean_metric(summary_df, dataset, metric):
    return float(summary_df.loc[summary_df["Dataset"] == dataset, f"{metric}_Mean"].iloc[0])


def save_excel_formatted(df, path, sheet_name="Sheet1"):
    ensure_output_dir(os.path.dirname(path))
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        ws = writer.sheets[sheet_name[:31]]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="center")
                cell.border = border
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 55)


def sci_axes(ax):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(axis="both", direction="out", length=3, width=0.8)
    ax.grid(False)


def save_figure(fig, out_dir, name):
    ensure_output_dir(out_dir)
    fig.savefig(os.path.join(out_dir, f"{name}.png"), dpi=600, bbox_inches="tight")
    fig.savefig(os.path.join(out_dir, f"{name}.pdf"), bbox_inches="tight")
    plt.close(fig)


def percent_ci_format(mean, lower, upper):
    return f"{100 * mean:.2f}% ({100 * lower:.2f}–{100 * upper:.2f}%)"




def warn_if_prezscore_input(file_path):
    """Warn if the input file name suggests that features were standardized before CV."""
    name = os.path.basename(str(file_path)).lower()
    if any(k in name for k in ["z-score", "z_scores", "z-scores", "zscore", "standard"]):
        print("警告：当前 DATA_FILE 文件名看起来像已经整体标准化过的数据。")
        print("为了完全避免标准化泄漏，建议改用未标准化的插补后原始训练集。")
        print("本脚本已把 StandardScaler 放进CV流程：每一折只在训练折fit，再transform验证折。")


def fit_transform_full_training_if_needed(X, output_dir=None, prefix="full_training"):
    """
    For post-CV final refitting / interpretation only.
    The scaler is fitted on the entire development cohort after model selection.
    Do NOT use external validation data here.
    """
    if not USE_SCALER:
        return X.copy()
    from sklearn.preprocessing import StandardScaler
    scaler = StandardScaler()
    X_scaled = pd.DataFrame(scaler.fit_transform(X), columns=X.columns, index=X.index)
    if output_dir is not None:
        try:
            scaler_df = pd.DataFrame({
                "Feature": X.columns,
                "Training_Mean": scaler.mean_,
                "Training_SD": scaler.scale_,
            })
            save_excel_formatted(
                scaler_df,
                os.path.join(output_dir, f"{prefix}_scaler_mean_sd.xlsx"),
                "Scaler_Mean_SD",
            )
        except Exception as e:
            print(f"Scaler mean/SD保存失败: {e}")
    return X_scaled


# =============================================================================
# 3. CV, plots, bootstrap
# =============================================================================
def run_one_setting(X, y, sampler_name, sampler_class, ratio):
    cv = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)
    fold_rows = []
    val_y_all, val_prob_all = [], []
    train_y_all, train_prob_all = [], []

    for fold_id, (train_idx, val_idx) in enumerate(cv.split(X, y), start=1):
        X_train, X_val = X.iloc[train_idx], X.iloc[val_idx]
        y_train, y_val = y.iloc[train_idx], y.iloc[val_idx]

        sampler = make_sampler(sampler_name, sampler_class, ratio, y_train)
        pipeline = make_pipeline(sampler, clone(build_model()))
        pipeline.fit(X_train, y_train)

        train_prob = get_positive_probability(pipeline, X_train)
        val_prob = get_positive_probability(pipeline, X_val)

        for dataset, yy, pp in [
            ("Train", y_train, train_prob),
            ("Validation", y_val, val_prob),
        ]:
            row = {
                "Fold": fold_id,
                "Dataset": dataset,
                "SMOTE_Method": sampler_name,
                "Sampling_Ratio": ratio,
            }
            row.update(calc_metrics(yy, pp, THRESHOLD))
            fold_rows.append(row)

        train_y_all.extend(y_train.tolist())
        train_prob_all.extend(train_prob.tolist())
        val_y_all.extend(y_val.tolist())
        val_prob_all.extend(val_prob.tolist())

    fold_df = pd.DataFrame(fold_rows)
    summary = summarize_fold_metrics(fold_df)
    val_auprc = get_mean_metric(summary, "Validation", "AUPRC")
    train_auprc = get_mean_metric(summary, "Train", "AUPRC")

    return {
        "Model": MODEL_SHORT,
        "SMOTE_Method": sampler_name,
        "Sampling_Ratio": ratio,
        "Train_AUPRC": train_auprc,
        "Val_AUPRC": val_auprc,
        "Gap": train_auprc - val_auprc,
        "Threshold": THRESHOLD,
        "Params": str(MODEL_PARAMS),
        "Fold_DF": fold_df,
        "Summary_DF": summary,
        "train_y": np.array(train_y_all),
        "train_prob": np.array(train_prob_all),
        "val_y": np.array(val_y_all),
        "val_prob": np.array(val_prob_all),
    }


def plot_combined_pr_roc(train_y, train_prob, val_y, val_prob, summary_df, result_row, output_dir):
    fig_dir = os.path.join(output_dir, "PR_ROC_curves_FINAL")
    ensure_output_dir(fig_dir)

    train_auprc_mean = float(result_row["Train_AUPRC"])
    val_auprc_mean = float(result_row["Val_AUPRC"])
    train_auc_mean = get_mean_metric(summary_df, "Train", "AUC")
    val_auc_mean = get_mean_metric(summary_df, "Validation", "AUC")

    # PR curve shape uses pooled CV predictions; legend uses five-fold mean, matching Excel.
    train_precision, train_recall, _ = precision_recall_curve(train_y, train_prob)
    val_precision, val_recall, _ = precision_recall_curve(val_y, val_prob)

    fig, ax = plt.subplots(figsize=(6, 5), dpi=600)
    ax.plot(train_recall, train_precision, linewidth=1.8, color=COLOR_TRAIN,
            label=f"Train mean AUPRC = {train_auprc_mean:.3f}")
    ax.plot(val_recall, val_precision, linewidth=1.8, color=COLOR_VAL,
            label=f"Validation mean AUPRC = {val_auprc_mean:.3f}")
    ax.set_xlabel("Recall")
    ax.set_ylabel("Precision")
    ax.set_title(f"{MODEL_SHORT} Precision-Recall Curve")
    ax.legend(frameon=False, loc="best")
    sci_axes(ax)
    fig.tight_layout()

    ratio_tag = str(result_row["Sampling_Ratio"]).replace(".", "p")
    val_tag = f"{val_auprc_mean:.3f}".replace(".", "p")
    method_tag = str(result_row["SMOTE_Method"])
    save_figure(fig, fig_dir, f"OPEN_THIS_PR_curve_BEST_{method_tag}_ratio_{ratio_tag}_ValAUPRC_{val_tag}")

    # ROC curve
    train_fpr, train_tpr, _ = roc_curve(train_y, train_prob)
    val_fpr, val_tpr, _ = roc_curve(val_y, val_prob)

    fig, ax = plt.subplots(figsize=(6, 5), dpi=600)
    ax.plot(train_fpr, train_tpr, linewidth=1.8, color=COLOR_TRAIN,
            label=f"Train mean AUC = {train_auc_mean:.3f}")
    ax.plot(val_fpr, val_tpr, linewidth=1.8, color=COLOR_VAL,
            label=f"Validation mean AUC = {val_auc_mean:.3f}")
    ax.plot([0, 1], [0, 1], linestyle="--", linewidth=1.0, color=COLOR_GREY)
    ax.set_xlabel("False Positive Rate")
    ax.set_ylabel("True Positive Rate")
    ax.set_title(f"{MODEL_SHORT} ROC Curve")
    ax.legend(frameon=False, loc="lower right")
    sci_axes(ax)
    fig.tight_layout()
    save_figure(fig, fig_dir, f"OPEN_THIS_ROC_curve_BEST_{method_tag}_ratio_{ratio_tag}")


def bootstrap_fold_metrics(fold_df, output_dir, best_model_name=MODEL_SHORT):
    """Bootstrap 95% CI from five-fold metric values after best setting selection."""
    rng = np.random.default_rng(BOOTSTRAP_RANDOM_STATE)
    metric_cols = [
        "AUPRC", "AUC", "Accuracy", "Balanced_Accuracy", "Recall_Sensitivity",
        "Specificity", "Precision", "F1", "MCC",
    ]

    records = []
    alpha = (100 - BOOTSTRAP_CI) / 2

    for dataset in ["Train", "Validation"]:
        sub = fold_df[fold_df["Dataset"] == dataset].reset_index(drop=True)
        n = len(sub)
        for metric in metric_cols:
            values = sub[metric].astype(float).values
            boot_means = []
            for _ in range(BOOTSTRAP_N):
                sample = rng.choice(values, size=n, replace=True)
                boot_means.append(np.nanmean(sample))
            lower, upper = np.nanpercentile(boot_means, [alpha, 100 - alpha])
            mean_value = np.nanmean(values)
            records.append({
                "Model": best_model_name,
                "Dataset": dataset,
                "Metric": metric,
                "Mean": float(mean_value),
                "SD": float(np.nanstd(values, ddof=1)) if n > 1 else 0.0,
                "CI_Lower": float(lower),
                "CI_Upper": float(upper),
                "Formatted_Percent_95CI": percent_ci_format(mean_value, lower, upper)
                if metric != "MCC" else f"{mean_value:.3f} ({lower:.3f}–{upper:.3f})",
                "N_Folds": n,
                "Bootstrap_N": BOOTSTRAP_N,
                "Threshold": THRESHOLD,
            })

    ci_df = pd.DataFrame(records)
    save_excel_formatted(ci_df, os.path.join(output_dir, "05_best_fold_bootstrap_95CI.xlsx"), "Bootstrap_95CI")

    val_ci = ci_df[ci_df["Dataset"] == "Validation"].set_index("Metric")
    paper_row = pd.DataFrame([{
        "Model": best_model_name,
        "Balanced accuracy": val_ci.loc["Balanced_Accuracy", "Formatted_Percent_95CI"],
        "Recall": val_ci.loc["Recall_Sensitivity", "Formatted_Percent_95CI"],
        "Precision": val_ci.loc["Precision", "Formatted_Percent_95CI"],
        "F1 score": val_ci.loc["F1", "Formatted_Percent_95CI"],
        "AUPRC": val_ci.loc["AUPRC", "Formatted_Percent_95CI"],
        "AUROC": val_ci.loc["AUC", "Formatted_Percent_95CI"],
        "Threshold": THRESHOLD,
        "Note": f"Values are bootstrap {BOOTSTRAP_CI}% confidence intervals based on five-fold cross-validation metrics.",
    }])
    save_excel_formatted(paper_row, os.path.join(output_dir, "06_paper_table_row_validation_95CI.xlsx"), "Paper_Table")

    return ci_df, paper_row


# =============================================================================
# 4. RFECV and selected-feature coefficients
# =============================================================================
def run_rfecv_and_plot_top10(X, y, sampler_name, sampler_class, ratio, output_dir):
    rfecv_dir = os.path.join(output_dir, "RFECV")
    ensure_output_dir(rfecv_dir)

    sampler = make_sampler(sampler_name, sampler_class, ratio, y)
    pipeline = make_pipeline(sampler, build_model())
    cv = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)

    rfecv = RFECV(
        estimator=pipeline,
        step=1,
        cv=cv,
        scoring="average_precision",
        n_jobs=-1,
        importance_getter="named_steps.model.coef_",
        min_features_to_select=1,
    )
    rfecv.fit(X, y)

    ranking = pd.DataFrame({
        "Feature": X.columns,
        "Selected_by_RFECV": rfecv.support_,
        "RFECV_Ranking": rfecv.ranking_,
    }).sort_values(["RFECV_Ranking", "Feature"], ascending=[True, True]).reset_index(drop=True)

    selected_features = X.columns[rfecv.support_].tolist()
    selected_df = pd.DataFrame({"Selected_Feature": selected_features})

    save_excel_formatted(selected_df, os.path.join(rfecv_dir, "RFECV_selected_features.xlsx"), "Selected")
    save_excel_formatted(ranking, os.path.join(rfecv_dir, "RFECV_feature_ranking.xlsx"), "RFECV_Ranking")

    cv_scores = pd.DataFrame({
        "Number_of_Selected_Features": np.arange(1, len(rfecv.cv_results_["mean_test_score"]) + 1),
        "Mean_CV_AUPRC": rfecv.cv_results_["mean_test_score"],
        "SD_CV_AUPRC": rfecv.cv_results_.get("std_test_score", np.nan),
    })
    save_excel_formatted(cv_scores, os.path.join(rfecv_dir, "RFECV_cv_scores.xlsx"), "CV_Scores")

    # RFECV-selected features -> resample -> final linear SVM -> absolute coefficients
    X_selected = X[selected_features]
    X_selected_model = fit_transform_full_training_if_needed(X_selected, rfecv_dir, prefix="rfecv_selected_full_training")
    final_sampler = make_sampler(sampler_name, sampler_class, ratio, y)
    X_res, y_res = final_sampler.fit_resample(X_selected_model, y)

    final_model = build_model()
    final_model.fit(X_res, y_res)

    coef = np.abs(final_model.coef_).ravel()
    coef_df = pd.DataFrame({
        "Feature": selected_features,
        "Abs_Coefficient_after_RFECV": coef,
    }).sort_values("Abs_Coefficient_after_RFECV", ascending=False).reset_index(drop=True)

    save_excel_formatted(
        coef_df,
        os.path.join(rfecv_dir, "RFECV_selected_abs_coefficients.xlsx"),
        "RFECV_Coefficients",
    )

    plot_df = coef_df.head(TOP_N_FEATURES).sort_values("Abs_Coefficient_after_RFECV", ascending=True)
    height = max(4.5, 0.38 * len(plot_df))
    fig, ax = plt.subplots(figsize=(8, height), dpi=600)
    ax.barh(
        plot_df["Feature"],
        plot_df["Abs_Coefficient_after_RFECV"],
        color=COLOR_BAR,
        edgecolor="black",
        linewidth=0.4,
    )
    ax.set_xlabel("Absolute coefficient after RFECV")
    ax.set_ylabel("")
    ax.set_title(f"Top {min(TOP_N_FEATURES, len(plot_df))} RFECV-Selected SVM Features")
    sci_axes(ax)
    fig.tight_layout()
    save_figure(fig, rfecv_dir, "OPEN_THIS_RFECV_top10_selected_features")

    return ranking, coef_df


# =============================================================================
# 5. Main
# =============================================================================
def main():
    ensure_output_dir(OUTPUT_DIR)
    X, y = load_data(DATA_FILE, TARGET_COL)
    warn_if_prezscore_input(DATA_FILE)

    print("=" * 70)
    print(f"N={len(y)}, Positive={int(y.sum())}, Negative={int(len(y) - y.sum())}")
    print(f"Positive ratio={y.mean():.4f}")
    print(f"Model params={MODEL_PARAMS}")
    print(f"SMOTE k_neighbors={SMOTE_K_NEIGHBORS}")
    print(f"Threshold={THRESHOLD}")
    print(f"Output folder={OUTPUT_DIR}")
    print("=" * 70)

    all_results = []
    all_folds = []
    summaries = {}
    full_objs = {}

    current_ratio = y.sum() / (len(y) - y.sum())

    for sampler_name, sampler_class in SAMPLER_CONFIGS.items():
        for ratio in SAMPLING_RATIOS:
            if ratio <= current_ratio:
                print(f"Skip {sampler_name} ratio={ratio}: ratio <= current minority/majority ratio")
                continue

            print(f"Running: {sampler_name}, ratio={ratio}")
            obj = run_one_setting(X, y, sampler_name, sampler_class, ratio)

            all_results.append({
                "Model": obj["Model"],
                "SMOTE_Method": obj["SMOTE_Method"],
                "Sampling_Ratio": obj["Sampling_Ratio"],
                "Train_AUPRC": obj["Train_AUPRC"],
                "Val_AUPRC": obj["Val_AUPRC"],
                "Gap": obj["Gap"],
                "Threshold": obj["Threshold"],
                "Params": obj["Params"],
            })
            all_folds.append(obj["Fold_DF"])
            key = f"{sampler_name}_{ratio}"
            summaries[key] = obj["Summary_DF"]
            full_objs[key] = obj

    if len(all_results) == 0:
        raise RuntimeError("没有任何SMOTE设置被运行。请检查 SAMPLING_RATIOS 是否大于当前少数类/多数类比例。")

    result_df = pd.DataFrame(all_results).sort_values(
        ["Val_AUPRC", "Gap"], ascending=[False, True]
    ).reset_index(drop=True)

    best = result_df.iloc[0]
    best_sampler_name = best["SMOTE_Method"]
    best_ratio = float(best["Sampling_Ratio"])
    best_key = f"{best_sampler_name}_{best_ratio}"
    best_sampler_class = SAMPLER_CONFIGS[best_sampler_name]
    best_obj = full_objs[best_key]
    best_summary = summaries[best_key]

    all_fold_df = pd.concat(all_folds, ignore_index=True)

    # Core Excel outputs
    save_excel_formatted(result_df, os.path.join(OUTPUT_DIR, "01_all_results.xlsx"), "All_Results")
    save_excel_formatted(all_fold_df, os.path.join(OUTPUT_DIR, "02_fold_metrics.xlsx"), "Fold_Metrics")
    save_excel_formatted(best_summary, os.path.join(OUTPUT_DIR, "03_best_summary_metrics_mean_sd.xlsx"), "Summary")
    save_excel_formatted(result_df.head(1), os.path.join(OUTPUT_DIR, "04_best_setting.xlsx"), "Best_Setting")

    # Check values: use this to verify figure legends.
    check_df = pd.DataFrame([{
        "Model": MODEL_SHORT,
        "Best_SMOTE_Method": best_sampler_name,
        "Best_Sampling_Ratio": best_ratio,
        "Train_AUPRC_for_PR_legend": float(best["Train_AUPRC"]),
        "Validation_AUPRC_for_PR_legend": float(best["Val_AUPRC"]),
        "Threshold": THRESHOLD,
        "Output_DIR": OUTPUT_DIR,
    }])
    save_excel_formatted(check_df, os.path.join(OUTPUT_DIR, "00_OPEN_THIS_CHECK_VALUES.xlsx"), "Check")

    # Combined PR/ROC curves: legend uses five-fold mean and matches Excel.
    plot_combined_pr_roc(
        best_obj["train_y"],
        best_obj["train_prob"],
        best_obj["val_y"],
        best_obj["val_prob"],
        best_summary,
        best,
        OUTPUT_DIR,
    )

    # Bootstrap 95% CI after the best setting is selected.
    best_fold_df = best_obj["Fold_DF"]
    ci_df, paper_row = bootstrap_fold_metrics(best_fold_df, OUTPUT_DIR, MODEL_SHORT)

    # RFECV + selected-feature coefficient plot.
    run_rfecv_and_plot_top10(X, y, best_sampler_name, best_sampler_class, best_ratio, OUTPUT_DIR)

    print("=" * 70)
    print("Finished.")
    print("IMPORTANT: Open this NEW output folder only:")
    print(OUTPUT_DIR)
    print("Best setting:")
    print(result_df.head(1).to_string(index=False))
    print("Best summary:")
    print(best_summary.to_string(index=False))
    print("Paper table row:")
    print(paper_row.to_string(index=False))
    print("=" * 70)


if __name__ == "__main__":
    main()
